"""Demo data for the „missing patient after import” troubleshooting video."""

from __future__ import annotations

from datetime import date
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]
FIXTURE_XLSX = (
    REPO_ROOT
    / "docs"
    / "manual"
    / "assets"
    / "fixtures"
    / "demo-doctolib-2-patients.xlsx"
)


def write_demo_doctolib_xlsx(
    path: Path,
    *,
    queue_date: date,
    standort_name: str,
    data_rows: list[tuple[str, str, str, str, str]],
) -> None:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws["A1"] = queue_date.strftime("%d.%m.%Y")
    ws["A2"] = f"Standort: {standort_name}"
    header_row = 4
    for col, title in enumerate(
        ("Vorname", "Nachname", "Geburtsdatum", "Telefon", "E-Mail"),
        start=1,
    ):
        ws.cell(header_row, col, title)
    r = header_row + 1
    for data in data_rows:
        for col, val in enumerate(data, start=1):
            ws.cell(r, col, val)
        r += 1
    wb.save(path)
    wb.close()


def seed_import_troubleshooting_demo(ctx: dict) -> None:
    """One patient in today's queue; second exists in DB but was not imported today."""
    from django.utils import timezone

    from apps.core.api_utils import assign_group_to_test_user
    from apps.reception.models import (
        ClinicSite,
        ConsultingRoom,
        DailyQueue,
        ImportSourceSystem,
        ImportStatus,
        ImportType,
        Patient,
        PatientImportBatch,
        QueueEntry,
        QueueEntryStatus,
        QueueStatus,
    )
    from apps.users.models import StaffUser

    pwd = ctx.get("password") or "ScreenshotDemo2026!"

    reception, _ = StaffUser.objects.get_or_create(
        username="screenshot_reception",
        defaults={
            "email": "screenshot_reception@example.invalid",
            "first_name": "Screenshot",
            "last_name": "screenshot_reception",
            "is_staff": True,
            "is_active": True,
        },
    )
    reception.set_password(pwd)
    reception.is_staff = True
    reception.is_active = True
    reception.save()
    reception.groups.clear()
    assign_group_to_test_user(reception, "Reception")

    clinic, _ = ClinicSite.objects.get_or_create(
        code="SCR",
        defaults={"name": "Screenshot Klinik Demo"},
    )
    room, _ = ConsultingRoom.objects.get_or_create(
        clinic_site=clinic,
        code="R-VID",
        defaults={"name": "Video Demo Raum"},
    )
    reception.clinic_sites.add(clinic)

    today = timezone.now().date()
    queue, _ = DailyQueue.objects.get_or_create(
        queue_date=today,
        clinic_site=clinic,
        consulting_room=room,
        defaults={
            "status": QueueStatus.OPEN,
            "created_by_user": reception,
            "shift_code": "FULL_DAY",
        },
    )
    queue.status = QueueStatus.OPEN
    queue.save(update_fields=["status", "updated_at"])

    QueueEntry.objects.filter(daily_queue=queue).delete()

    p_imported, _ = Patient.objects.update_or_create(
        email="erika.mustermann.demo@example.invalid",
        defaults={
            "first_name": "Erika",
            "last_name": "Mustermann",
            "date_of_birth": date(1982, 3, 12),
            "phone": "4915111111111",
        },
    )
    p_imported.clinic_sites.add(clinic)

    p_missing, _ = Patient.objects.update_or_create(
        email="thomas.schneider.demo@example.invalid",
        defaults={
            "first_name": "Thomas",
            "last_name": "Schneider",
            "date_of_birth": date(1975, 11, 8),
            "phone": "4915222222222",
        },
    )
    p_missing.clinic_sites.add(clinic)

    entry = QueueEntry.objects.create(
        daily_queue=queue,
        patient=p_imported,
        position_no=1,
        entry_status=QueueEntryStatus.WAITING,
        created_by_user=reception,
    )

    write_demo_doctolib_xlsx(
        FIXTURE_XLSX,
        queue_date=today,
        standort_name=clinic.name,
        data_rows=[
            (
                "Erika",
                "Mustermann",
                "12.03.1982",
                "+49 151 11111111",
                "erika.mustermann.demo@example.invalid",
            ),
            (
                "Thomas",
                "Schneider",
                "08.11.1975",
                "+49 152 22222222",
                "thomas.schneider.demo@example.invalid",
            ),
        ],
    )

    batch_name = f"doctolib-pro-{today.isoformat()}T10-00-00-demo.xlsx"
    PatientImportBatch.objects.filter(source_file_name=batch_name).delete()
    batch = PatientImportBatch.objects.create(
        source_file_name=batch_name,
        source_file_sha256="demo" + "0" * 60,
        import_type=ImportType.DAILY_FILE_IMPORT,
        source_system=ImportSourceSystem.DOCTOLIB_EXPORT,
        status=ImportStatus.COMPLETED,
        total_rows=1,
        inserted_rows=1,
        matched_rows=0,
        skipped_already_present_count=0,
        error_rows=0,
        created_by_user=reception,
        finished_at=timezone.now(),
    )

    ctx.update(
        {
            "password": pwd,
            "reception": reception,
            "clinic": clinic,
            "queue": queue,
            "queue_id": str(queue.id),
            "queue_date": today.isoformat(),
            "patient_imported_id": str(p_imported.id),
            "patient_missing_id": str(p_missing.id),
            "patient_missing_last_name": p_missing.last_name,
            "import_batch_id": str(batch.id),
            "queue_entry_imported_id": str(entry.id),
            "fixture_xlsx": str(FIXTURE_XLSX),
        }
    )


def seed_ctx_json(path: Path) -> dict:
    """Serializable subset of ctx for Playwright-only runs on the host."""
    ctx: dict = {}
    seed_import_troubleshooting_demo(ctx)
    payload = {
        "password": ctx["password"],
        "queue_id": ctx["queue_id"],
        "queue_date": ctx["queue_date"],
        "patient_missing_id": ctx["patient_missing_id"],
        "patient_missing_last_name": ctx["patient_missing_last_name"],
        "import_batch_id": ctx["import_batch_id"],
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(__import__("json").dumps(payload, indent=2), encoding="utf-8")
    return payload


if __name__ == "__main__":
    import json
    import sys

    repo = Path(__file__).resolve().parents[2]
    if str(repo) not in sys.path:
        sys.path.insert(0, str(repo))

    from scripts.manual_demo import setup_django

    out = repo / "docs" / "manual" / "_build" / "import-troubleshooting-ctx.json"
    setup_django()
    data = seed_ctx_json(out)
    print(json.dumps(data, indent=2))
    print(f"Wrote {out}")
