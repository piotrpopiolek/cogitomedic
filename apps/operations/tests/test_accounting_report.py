"""Accounting weekly report service, export, and admin views."""

from __future__ import annotations

import uuid
from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo

from io import BytesIO

from django.contrib.auth.models import Group
from django.test import Client, TestCase
from django.urls import reverse
from django.utils import timezone

from openpyxl import load_workbook
from pydantic import ValidationError

from apps.core.api_utils import assign_group_to_test_user
from apps.intake.models import IntakeStatus, PatientIntakeForm
from apps.medical.models import (
    DocVersionStatus,
    MedicalDocStatus,
    MedicalDocument,
    MedicalDocumentSourceType,
    MedicalDocumentVersion,
    PdfStatus,
)
from apps.operations.accounting_report import (
    AccountingReportResult,
    AccountingReportRow,
    DoctorPublicationCount,
    build_accounting_report,
    default_report_week_range,
    format_patient_address,
    format_patient_postal_city,
    format_patient_street,
    normalize_postal_code_display,
    published_at_range_utc,
    resolve_report_date_range,
)
from apps.operations.api_schemas import (
    AccountingReportQueryParams,
    build_accounting_report_response,
)
from apps.operations.export import (
    render_accounting_report_csv,
    render_accounting_report_xlsx,
)
from apps.operations.models import AuditEvent
from apps.operations.accounting_access import accounting_report_access_ok
from apps.reception.models import (
    ClinicSite,
    ConsultingRoom,
    DailyQueue,
    Patient,
    PatientFormSession,
    QueueEntry,
    QueueEntryStatus,
    QueueStatus,
)
from apps.users.models import ROLE_GROUP_NAME_MAP, StaffUser


class AccountingReportBase(TestCase):
    @classmethod
    def setUpTestData(cls) -> None:
        cls.doctor = StaffUser.objects.create_user(
            username="acct-doctor",
            email="acct-doctor@example.com",
            password="test-pass-123",
            first_name="Hans",
            last_name="Müller",
            is_staff=True,
        )
        Group.objects.get_or_create(name=ROLE_GROUP_NAME_MAP["DOCTOR"])[0].user_set.add(
            cls.doctor
        )
        cls.doctor2 = StaffUser.objects.create_user(
            username="acct-doctor-2",
            email="acct-doctor-2@example.com",
            password="test-pass-123",
            first_name="Eva",
            last_name="Schmidt",
            is_staff=True,
        )
        Group.objects.get_or_create(name=ROLE_GROUP_NAME_MAP["DOCTOR"])[0].user_set.add(
            cls.doctor2
        )
        cls.accounting_user = StaffUser.objects.create_user(
            username="acct-user",
            email="acct-user@example.com",
            password="test-pass-123",
            is_staff=True,
        )
        Group.objects.get_or_create(name=ROLE_GROUP_NAME_MAP["ACCOUNTING"])[
            0
        ].user_set.add(cls.accounting_user)
        cls.clinic_site = ClinicSite.objects.create(
            code="ACCT",
            name="Accounting Clinic",
        )
        cls.consulting_room = ConsultingRoom.objects.create(
            clinic_site=cls.clinic_site,
            code="A1",
            name="Room A1",
        )
        cls.daily_queue = DailyQueue.objects.create(
            clinic_site=cls.clinic_site,
            consulting_room=cls.consulting_room,
            queue_date=date(2026, 3, 10),
            status=QueueStatus.OPEN,
            assigned_doctor=cls.doctor,
            created_by_user=cls.doctor,
        )
        cls.patient = Patient.objects.create(
            first_name="Anna",
            last_name="Kowalska",
            date_of_birth=date(1985, 5, 15),
            phone="48500111222",
            email="anna@example.com",
            street="Musterstr. 1",
            postal_code="10115",
            city="Berlin",
        )
        cls.queue_entry = QueueEntry.objects.create(
            daily_queue=cls.daily_queue,
            patient=cls.patient,
            entry_status=QueueEntryStatus.PATIENT_COMPLETED,
            position_no=1,
            created_by_user=cls.doctor,
        )
        cls.session = PatientFormSession.objects.create(
            queue_entry=cls.queue_entry,
            form_locale="de-DE",
            expires_at=timezone.now() + timedelta(hours=1),
            created_by_user=cls.doctor,
        )
        cls.intake = PatientIntakeForm.objects.create(
            queue_entry=cls.queue_entry,
            session=cls.session,
            form_status=IntakeStatus.SUBMITTED,
            submitted_at=timezone.now(),
            signature_sha256="a" * 64,
        )

    def _make_doc(self, **overrides) -> MedicalDocument:
        defaults = dict(
            queue_entry=self.queue_entry,
            intake_form=self.intake,
            status=MedicalDocStatus.PUBLISHED,
            current_version_no=1,
            created_by_user=self.doctor,
        )
        defaults.update(overrides)
        return MedicalDocument.objects.create(**defaults)

    def _make_published_version(
        self,
        doc: MedicalDocument,
        *,
        version_no: int = 1,
        published_at: datetime | None = None,
        published_by_user: StaffUser | None = None,
        revoked_at: datetime | None = None,
    ) -> MedicalDocumentVersion:
        when = published_at or timezone.now()
        version = MedicalDocumentVersion.objects.create(
            medical_document=doc,
            version_no=version_no,
            version_status=DocVersionStatus.PUBLISHED,
            pdf_generation_status=PdfStatus.COMPLETED,
            medical_payload_schema_version=1,
            medical_payload={"schema_version": 1},
            pdf_local_path="/media/befund/test.pdf",
            publish_request_id=uuid.uuid4(),
            published_at=when,
            publish_locale="de-DE",
            published_by_user=published_by_user or self.doctor,
            revoked_at=revoked_at,
        )
        MedicalDocument.objects.filter(pk=doc.pk).update(
            published_version_no=version_no,
            current_version_no=version_no,
        )
        return version

    def _create_publication_at_clinic(
        self,
        *,
        clinic_site: ClinicSite,
        queue_date: date,
        position_no: int,
        patient_phone: str,
        published_at: datetime,
    ) -> MedicalDocument:
        room = ConsultingRoom.objects.create(
            clinic_site=clinic_site,
            code=f"R{position_no}",
            name=f"Room {position_no}",
        )
        daily_queue = DailyQueue.objects.create(
            clinic_site=clinic_site,
            consulting_room=room,
            queue_date=queue_date,
            status=QueueStatus.OPEN,
            assigned_doctor=self.doctor,
            created_by_user=self.doctor,
        )
        patient = Patient.objects.create(
            first_name="Scope",
            last_name=f"Patient{position_no}",
            date_of_birth=date(1988, 1, 1),
            phone=patient_phone,
            email=f"scope{position_no}@example.com",
        )
        entry = QueueEntry.objects.create(
            daily_queue=daily_queue,
            patient=patient,
            entry_status=QueueEntryStatus.PATIENT_COMPLETED,
            position_no=position_no,
            created_by_user=self.doctor,
        )
        session = PatientFormSession.objects.create(
            queue_entry=entry,
            form_locale="de-DE",
            expires_at=timezone.now() + timedelta(hours=1),
            created_by_user=self.doctor,
        )
        intake = PatientIntakeForm.objects.create(
            queue_entry=entry,
            session=session,
            form_status=IntakeStatus.SUBMITTED,
            submitted_at=timezone.now(),
            signature_sha256="d" * 64,
        )
        doc = MedicalDocument.objects.create(
            queue_entry=entry,
            intake_form=intake,
            status=MedicalDocStatus.PUBLISHED,
            current_version_no=1,
            created_by_user=self.doctor,
        )
        self._make_published_version(doc, published_at=published_at)
        return doc


class AccountingReportServiceTests(AccountingReportBase):
    def test_includes_first_publication_in_range(self) -> None:
        doc = self._make_doc()
        published_at = datetime(2026, 3, 11, 10, 0, tzinfo=ZoneInfo("Europe/Warsaw"))
        self._make_published_version(doc, published_at=published_at)
        report = build_accounting_report(
            date_from=date(2026, 3, 10),
            date_to=date(2026, 3, 16),
        )
        self.assertEqual(len(report.rows), 1)
        row = report.rows[0]
        self.assertEqual(row.row_no, 1)
        self.assertEqual(row.first_name, "Anna")
        self.assertEqual(row.last_name, "Kowalska")
        self.assertEqual(row.street, "Musterstr. 1")
        self.assertEqual(row.postal_city, "10115 Berlin")
        self.assertEqual(row.email, "anna@example.com")
        self.assertEqual(row.doctor_name, "Hans Müller")
        self.assertEqual(row.exam_date, "10.03.2026")

    def test_revision_v2_not_in_report(self) -> None:
        doc = self._make_doc()
        v1_at = datetime(2026, 3, 11, 10, 0, tzinfo=ZoneInfo("Europe/Warsaw"))
        v2_at = datetime(2026, 3, 12, 10, 0, tzinfo=ZoneInfo("Europe/Warsaw"))
        self._make_published_version(doc, version_no=1, published_at=v1_at)
        self._make_published_version(
            doc, version_no=2, published_at=v2_at, published_by_user=self.doctor2
        )
        report = build_accounting_report(
            date_from=date(2026, 3, 10),
            date_to=date(2026, 3, 16),
        )
        self.assertEqual(len(report.rows), 1)
        self.assertEqual(report.rows[0].doctor_name, "Hans Müller")

    def test_external_upload_excluded(self) -> None:
        doc = self._make_doc(source_type=MedicalDocumentSourceType.EXTERNAL_UPLOAD)
        published_at = datetime(2026, 3, 11, 10, 0, tzinfo=ZoneInfo("Europe/Warsaw"))
        self._make_published_version(doc, published_at=published_at)
        report = build_accounting_report(
            date_from=date(2026, 3, 10),
            date_to=date(2026, 3, 16),
        )
        self.assertEqual(report.rows, [])

    def test_out_of_range_excluded(self) -> None:
        doc = self._make_doc()
        published_at = datetime(2026, 2, 1, 10, 0, tzinfo=ZoneInfo("Europe/Warsaw"))
        self._make_published_version(doc, published_at=published_at)
        report = build_accounting_report(
            date_from=date(2026, 3, 10),
            date_to=date(2026, 3, 16),
        )
        self.assertEqual(report.rows, [])

    def test_doctor_counts_aggregate_documents(self) -> None:
        doc1 = self._make_doc()
        self._make_published_version(
            doc1,
            published_at=datetime(2026, 3, 11, 10, 0, tzinfo=ZoneInfo("Europe/Warsaw")),
        )
        other_entry = QueueEntry.objects.create(
            daily_queue=self.daily_queue,
            patient=Patient.objects.create(
                first_name="Jan",
                last_name="Nowak",
                date_of_birth=date(1990, 1, 1),
                phone="48500999888",
                email="jan@example.com",
            ),
            entry_status=QueueEntryStatus.PATIENT_COMPLETED,
            position_no=2,
            created_by_user=self.doctor,
        )
        other_session = PatientFormSession.objects.create(
            queue_entry=other_entry,
            form_locale="de-DE",
            expires_at=timezone.now() + timedelta(hours=1),
            created_by_user=self.doctor,
        )
        other_intake = PatientIntakeForm.objects.create(
            queue_entry=other_entry,
            session=other_session,
            form_status=IntakeStatus.SUBMITTED,
            submitted_at=timezone.now(),
            signature_sha256="b" * 64,
        )
        doc2 = MedicalDocument.objects.create(
            queue_entry=other_entry,
            intake_form=other_intake,
            status=MedicalDocStatus.PUBLISHED,
            current_version_no=1,
            created_by_user=self.doctor2,
        )
        self._make_published_version(
            doc2,
            published_at=datetime(2026, 3, 12, 10, 0, tzinfo=ZoneInfo("Europe/Warsaw")),
            published_by_user=self.doctor2,
        )
        report = build_accounting_report(
            date_from=date(2026, 3, 10),
            date_to=date(2026, 3, 16),
        )
        counts = {item.doctor_name: item.count for item in report.doctor_counts}
        self.assertEqual(counts["Hans Müller"], 1)
        self.assertEqual(counts["Eva Schmidt"], 1)

    def test_empty_address_parts_return_empty_string(self) -> None:
        patient = Patient.objects.create(
            first_name="No",
            last_name="Address",
            date_of_birth=date(1991, 1, 1),
            phone="48500777666",
            email="noaddr@example.com",
        )
        self.assertEqual(format_patient_street(patient), "")
        self.assertEqual(format_patient_postal_city(patient), "")
        self.assertEqual(format_patient_address(patient), "")

    def test_postal_city_formats_plz_and_ort(self) -> None:
        patient = Patient.objects.create(
            first_name="Max",
            last_name="Mustermann",
            date_of_birth=date(1990, 1, 1),
            phone="48500777667",
            email="max@example.com",
            postal_code="10115",
            city="Berlin",
        )
        self.assertEqual(format_patient_postal_city(patient), "10115 Berlin")

    def test_postal_city_partial_fields(self) -> None:
        patient = Patient.objects.create(
            first_name="Eva",
            last_name="Test",
            date_of_birth=date(1990, 1, 1),
            phone="48500777668",
            email="eva@example.com",
            postal_code="10115",
        )
        self.assertEqual(format_patient_postal_city(patient), "10115")

    def test_postal_code_strips_excel_float_artifact(self) -> None:
        self.assertEqual(normalize_postal_code_display("17498.0"), "17498")
        patient = Patient.objects.create(
            first_name="PLZ",
            last_name="Fix",
            date_of_birth=date(1990, 1, 1),
            phone="48500777669",
            email="plz@example.com",
            postal_code="17498.0",
            city="Neubrandenburg",
        )
        self.assertEqual(format_patient_postal_city(patient), "17498 Neubrandenburg")

    def test_default_week_range_is_monday_to_sunday(self) -> None:
        monday, sunday = default_report_week_range(today=date(2026, 3, 11))
        self.assertEqual(monday, date(2026, 3, 9))
        self.assertEqual(sunday, date(2026, 3, 15))

    def test_resolve_report_date_range_swaps_inverted(self) -> None:
        date_from, date_to = resolve_report_date_range(
            date_from_raw="2026-03-20",
            date_to_raw="2026-03-10",
        )
        self.assertEqual(date_from, date(2026, 3, 10))
        self.assertEqual(date_to, date(2026, 3, 20))

    def test_published_at_range_utc_respects_local_midnight(self) -> None:
        start, end = published_at_range_utc(date(2026, 3, 10), date(2026, 3, 10))
        self.assertLess(start, end)
        self.assertEqual((end - start).days, 1)

    def test_revoked_publication_excluded(self) -> None:
        doc = self._make_doc()
        published_at = datetime(2026, 3, 11, 10, 0, tzinfo=ZoneInfo("Europe/Warsaw"))
        self._make_published_version(
            doc,
            published_at=published_at,
            revoked_at=published_at + timedelta(hours=1),
        )
        report = build_accounting_report(
            date_from=date(2026, 3, 10),
            date_to=date(2026, 3, 16),
        )
        self.assertEqual(report.rows, [])

    def test_manager_scope_limits_rows_to_assigned_clinic_sites(self) -> None:
        published_at = datetime(2026, 3, 11, 10, 0, tzinfo=ZoneInfo("Europe/Warsaw"))
        self._create_publication_at_clinic(
            clinic_site=self.clinic_site,
            queue_date=date(2026, 3, 10),
            position_no=10,
            patient_phone="48500111001",
            published_at=published_at,
        )
        other_site = ClinicSite.objects.create(code="OTH", name="Other Clinic")
        self._create_publication_at_clinic(
            clinic_site=other_site,
            queue_date=date(2026, 3, 10),
            position_no=11,
            patient_phone="48500111002",
            published_at=published_at,
        )
        report = build_accounting_report(
            date_from=date(2026, 3, 10),
            date_to=date(2026, 3, 16),
            scoped_clinic_site_ids=[self.clinic_site.id],
        )
        self.assertEqual(len(report.rows), 1)
        self.assertEqual(report.rows[0].last_name, "Patient10")


class AccountingReportExportTests(AccountingReportBase):
    def test_csv_contains_german_headers(self) -> None:
        doc = self._make_doc()
        self._make_published_version(
            doc,
            published_at=datetime(2026, 3, 11, 10, 0, tzinfo=ZoneInfo("Europe/Warsaw")),
        )
        report = build_accounting_report(
            date_from=date(2026, 3, 10),
            date_to=date(2026, 3, 16),
        )
        content = render_accounting_report_csv(report.rows).decode("utf-8")
        self.assertIn("Vorname", content)
        self.assertIn("Straße", content)
        self.assertIn("PLZ/Ort", content)
        self.assertIn("Anna", content)
        self.assertIn("10.03.2026", content)

    def test_xlsx_export_contains_data_rows(self) -> None:
        doc = self._make_doc()
        self._make_published_version(
            doc,
            published_at=datetime(2026, 3, 11, 10, 0, tzinfo=ZoneInfo("Europe/Warsaw")),
        )
        report = build_accounting_report(
            date_from=date(2026, 3, 10),
            date_to=date(2026, 3, 16),
        )
        content = render_accounting_report_xlsx(report.rows)
        workbook = load_workbook(BytesIO(content))
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(rows[0][1], "Vorname")
        self.assertIn("Anna", rows[1])


class AccountingReportViewTests(AccountingReportBase):
    def setUp(self) -> None:
        self.client = Client()

    def test_accounting_user_can_open_dashboard(self) -> None:
        self.client.force_login(self.accounting_user)
        response = self.client.get(reverse("admin_accounting_report"))
        self.assertEqual(response.status_code, 200)

    def test_dashboard_includes_auto_submit_form_markup(self) -> None:
        self.client.force_login(self.accounting_user)
        response = self.client.get(reverse("admin_accounting_report"))
        content = response.content.decode()
        self.assertIn('id="accounting-report-form"', content)
        self.assertIn("accounting-report-form.js", content)
        self.assertIn("data-export-csv-url", content)
        self.assertIn("data-export-csv", content)
        self.assertIn("data-export-xlsx", content)

    def test_export_querystring_omits_page_and_keeps_dates(self) -> None:
        self.client.force_login(self.accounting_user)
        response = self.client.get(
            reverse("admin_accounting_report"),
            {
                "date_from": "2026-03-10",
                "date_to": "2026-03-16",
                "page": "2",
                "page_size": "20",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.context["export_querystring"],
            "?date_from=2026-03-10&date_to=2026-03-16",
        )

    def test_date_range_get_without_page_returns_first_page(self) -> None:
        doc = self._make_doc()
        self._make_published_version(
            doc,
            published_at=datetime(2026, 3, 11, 10, 0, tzinfo=ZoneInfo("Europe/Warsaw")),
        )
        self.client.force_login(self.accounting_user)
        response = self.client.get(
            reverse("admin_accounting_report"),
            {"date_from": "2026-03-10", "date_to": "2026-03-16"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["pagination"]["page"], 1)
        self.assertEqual(response.context["pagination"]["total"], 1)

    def test_dashboard_includes_dark_mode_table_classes(self) -> None:
        self.client.force_login(self.accounting_user)
        response = self.client.get(reverse("admin_accounting_report"))
        content = response.content.decode()
        self.assertIn("dark:text-base-100", content)
        self.assertIn("dark:bg-base-900", content)
        self.assertIn("dark:bg-base-800", content)

    def test_accounting_user_admin_index_redirects_to_report(self) -> None:
        self.client.force_login(self.accounting_user)
        response = self.client.get(reverse("admin:index"))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], reverse("admin_accounting_report"))

    def test_admin_user_admin_index_stays_on_index(self) -> None:
        admin = StaffUser.objects.create_user(
            username="acct-admin-index",
            email="acct-admin-index@example.com",
            password="test-pass-123",
            is_staff=True,
        )
        assign_group_to_test_user(admin, "Admin")
        self.client.force_login(admin)
        response = self.client.get(reverse("admin:index"))
        self.assertEqual(response.status_code, 200)

    def test_admin_user_can_open_dashboard(self) -> None:
        admin = StaffUser.objects.create_user(
            username="acct-admin",
            email="acct-admin@example.com",
            password="test-pass-123",
            is_staff=True,
        )
        assign_group_to_test_user(admin, "Admin")
        self.client.force_login(admin)
        response = self.client.get(reverse("admin_accounting_report"))
        self.assertEqual(response.status_code, 200)

    def test_manager_sees_only_assigned_clinic_in_dashboard(self) -> None:
        published_at = datetime(2026, 3, 11, 10, 0, tzinfo=ZoneInfo("Europe/Warsaw"))
        self._create_publication_at_clinic(
            clinic_site=self.clinic_site,
            queue_date=date(2026, 3, 10),
            position_no=20,
            patient_phone="48500222001",
            published_at=published_at,
        )
        other_site = ClinicSite.objects.create(code="MGR", name="Manager Other Clinic")
        self._create_publication_at_clinic(
            clinic_site=other_site,
            queue_date=date(2026, 3, 10),
            position_no=21,
            patient_phone="48500222002",
            published_at=published_at,
        )
        manager = StaffUser.objects.create_user(
            username="acct-manager",
            email="acct-manager@example.com",
            password="test-pass-123",
            is_staff=True,
        )
        assign_group_to_test_user(manager, "Manager")
        manager.clinic_sites.add(self.clinic_site)
        self.client.force_login(manager)
        response = self.client.get(
            reverse("admin_accounting_report"),
            {"date_from": "2026-03-10", "date_to": "2026-03-16"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["pagination"]["total"], 1)
        self.assertEqual(response.context["items"][0].last_name, "Patient20")

    def test_doctor_forbidden(self) -> None:
        self.client.force_login(self.doctor)
        response = self.client.get(reverse("admin_accounting_report"))
        self.assertEqual(response.status_code, 403)

    def test_reception_forbidden(self) -> None:
        reception = StaffUser.objects.create_user(
            username="acct-reception",
            email="acct-reception@example.com",
            password="test-pass-123",
            is_staff=True,
        )
        assign_group_to_test_user(reception, "Reception")
        self.client.force_login(reception)
        response = self.client.get(reverse("admin_accounting_report"))
        self.assertEqual(response.status_code, 403)

    def test_tablet_forbidden(self) -> None:
        tablet = StaffUser.objects.create_user(
            username="acct-tablet",
            email="acct-tablet@example.com",
            password="test-pass-123",
            is_staff=True,
        )
        assign_group_to_test_user(tablet, "Tablet")
        self.client.force_login(tablet)
        response = self.client.get(reverse("admin_accounting_report"))
        self.assertEqual(response.status_code, 403)

    def test_accounting_user_cannot_open_patient_admin_changelist(self) -> None:
        self.client.force_login(self.accounting_user)
        response = self.client.get(reverse("admin:reception_patient_changelist"))
        self.assertEqual(response.status_code, 403)

    def test_access_helper(self) -> None:
        self.assertTrue(accounting_report_access_ok(self.accounting_user))
        self.assertFalse(accounting_report_access_ok(self.doctor))

    def test_export_csv_writes_audit_event(self) -> None:
        doc = self._make_doc()
        self._make_published_version(
            doc,
            published_at=datetime(2026, 3, 11, 10, 0, tzinfo=ZoneInfo("Europe/Warsaw")),
        )
        self.client.force_login(self.accounting_user)
        response = self.client.get(
            reverse("admin_accounting_report_export_csv"),
            {
                "date_from": "2026-03-10",
                "date_to": "2026-03-16",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv; charset=utf-8")
        ev = AuditEvent.objects.filter(event_type="ACCOUNTING_REPORT_EXPORT").first()
        self.assertIsNotNone(ev)
        assert ev is not None
        self.assertEqual(ev.metadata.get("format"), "csv")
        self.assertEqual(ev.metadata.get("row_count"), 1)

    def test_export_xlsx_writes_audit_event(self) -> None:
        doc = self._make_doc()
        self._make_published_version(
            doc,
            published_at=datetime(2026, 3, 11, 10, 0, tzinfo=ZoneInfo("Europe/Warsaw")),
        )
        self.client.force_login(self.accounting_user)
        response = self.client.get(
            reverse("admin_accounting_report_export_xlsx"),
            {
                "date_from": "2026-03-10",
                "date_to": "2026-03-16",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertIn("Anna", rows[1])
        ev = AuditEvent.objects.filter(
            event_type="ACCOUNTING_REPORT_EXPORT",
            metadata__format="xlsx",
        ).first()
        self.assertIsNotNone(ev)
        assert ev is not None
        self.assertEqual(ev.metadata.get("row_count"), 1)

    def test_pagination_defaults_to_fifty_rows(self) -> None:
        for index in range(25):
            patient = Patient.objects.create(
                first_name=f"P{index}",
                last_name="Test",
                date_of_birth=date(1980, 1, 1),
                phone=f"48500{index:06d}",
                email=f"p{index}@example.com",
            )
            entry = QueueEntry.objects.create(
                daily_queue=self.daily_queue,
                patient=patient,
                entry_status=QueueEntryStatus.PATIENT_COMPLETED,
                position_no=10 + index,
                created_by_user=self.doctor,
            )
            session = PatientFormSession.objects.create(
                queue_entry=entry,
                form_locale="de-DE",
                expires_at=timezone.now() + timedelta(hours=1),
                created_by_user=self.doctor,
            )
            intake = PatientIntakeForm.objects.create(
                queue_entry=entry,
                session=session,
                form_status=IntakeStatus.SUBMITTED,
                submitted_at=timezone.now(),
                signature_sha256="c" * 64,
            )
            doc = MedicalDocument.objects.create(
                queue_entry=entry,
                intake_form=intake,
                status=MedicalDocStatus.PUBLISHED,
                current_version_no=1,
                created_by_user=self.doctor,
            )
            self._make_published_version(
                doc,
                published_at=datetime(
                    2026, 3, 11, 9, index, tzinfo=ZoneInfo("Europe/Warsaw")
                ),
            )
        self.client.force_login(self.accounting_user)
        response = self.client.get(
            reverse("admin_accounting_report"),
            {"date_from": "2026-03-10", "date_to": "2026-03-16"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["items"]), 25)
        self.assertEqual(response.context["pagination"]["total"], 25)
        self.assertEqual(response.context["pagination"]["page_size"], 50)

    def test_pagination_rejects_invalid_page_size(self) -> None:
        self.client.force_login(self.accounting_user)
        response = self.client.get(
            reverse("admin_accounting_report"),
            {"date_from": "2026-03-10", "date_to": "2026-03-16", "page_size": "25"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["pagination"]["page_size"], 50)

    def test_query_params_defaults_and_resolved_week(self) -> None:
        query = AccountingReportQueryParams.model_validate({})
        self.assertEqual(query.page, 1)
        self.assertEqual(query.page_size, 50)
        date_from, date_to = query.resolved_date_range()
        expected_from, expected_to = default_report_week_range()
        self.assertEqual(date_from, expected_from)
        self.assertEqual(date_to, expected_to)

    def test_query_params_parses_iso_dates(self) -> None:
        query = AccountingReportQueryParams.model_validate(
            {
                "date_from": "2026-03-10",
                "date_to": "2026-03-16",
                "page": "2",
                "page_size": "10",
            }
        )
        self.assertEqual(query.date_from, date(2026, 3, 10))
        self.assertEqual(query.date_to, date(2026, 3, 16))
        self.assertEqual(query.page, 2)
        self.assertEqual(query.page_size, 10)

    def test_query_params_rejects_invalid_page_size(self) -> None:
        with self.assertRaises(ValidationError):
            AccountingReportQueryParams.model_validate({"page_size": "25"})

    def test_query_params_rejects_invalid_date(self) -> None:
        with self.assertRaises(ValidationError):
            AccountingReportQueryParams.model_validate({"date_from": "not-a-date"})

    def test_query_params_rejects_extra_fields(self) -> None:
        with self.assertRaises(ValidationError):
            AccountingReportQueryParams.model_validate({"unexpected": "x"})

    def test_build_response_serializes_report_with_pagination(self) -> None:
        doc_id = uuid.uuid4()
        doctor_id = self.doctor.id
        report = AccountingReportResult(
            rows=[
                AccountingReportRow(
                    row_no=index,
                    first_name=f"First{index}",
                    last_name=f"Last{index}",
                    street="Street 1",
                    postal_city="10115 Berlin",
                    email=f"p{index}@example.com",
                    doctor_name="Dr. Test",
                    exam_date="10.03.2026",
                    medical_document_id=doc_id,
                    doctor_user_id=doctor_id,
                )
                for index in range(1, 4)
            ],
            doctor_counts=[
                DoctorPublicationCount(
                    doctor_user_id=doctor_id,
                    doctor_name="Dr. Test",
                    count=3,
                )
            ],
            date_from=date(2026, 3, 10),
            date_to=date(2026, 3, 16),
        )
        response = build_accounting_report_response(report, page=2, page_size=2)
        payload = response.model_dump(mode="json")
        self.assertEqual(payload["date_from"], "2026-03-10")
        self.assertEqual(payload["date_to"], "2026-03-16")
        self.assertEqual(payload["report_total_rows"], 3)
        self.assertEqual(payload["pagination"], {"page": 2, "page_size": 2, "total": 3})
        self.assertEqual(len(payload["items"]), 1)
        self.assertEqual(payload["items"][0]["row_no"], 3)
        self.assertEqual(payload["items"][0]["medical_document_id"], str(doc_id))
        self.assertEqual(payload["doctor_counts"][0]["count"], 3)

    def test_build_response_matches_service_output(self) -> None:
        doc = self._make_doc()
        self._make_published_version(
            doc,
            published_at=datetime(2026, 3, 11, 10, 0, tzinfo=ZoneInfo("Europe/Warsaw")),
        )
        report = build_accounting_report(
            date_from=date(2026, 3, 10),
            date_to=date(2026, 3, 16),
        )
        query = AccountingReportQueryParams.model_validate(
            {"date_from": "2026-03-10", "date_to": "2026-03-16"}
        )
        date_from, date_to = query.resolved_date_range()
        self.assertEqual(report.date_from, date_from)
        self.assertEqual(report.date_to, date_to)
        response = build_accounting_report_response(
            report,
            page=query.page,
            page_size=query.page_size,
        )
        self.assertEqual(len(response.items), 1)
        self.assertEqual(response.items[0].first_name, "Anna")
        self.assertEqual(response.items[0].last_name, "Kowalska")
