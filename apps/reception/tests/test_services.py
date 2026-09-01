from __future__ import annotations

import importlib
import tempfile
import uuid
from datetime import date, timedelta
from pathlib import Path
from unittest.mock import patch

from django.contrib.admin.sites import AdminSite
from django.apps import apps as django_apps
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from django.db.models import Q
from django.test import Client, RequestFactory, TestCase
from django.urls import reverse
from django.utils import timezone

from apps.intake.models import IntakeStatus, PatientIntakeForm
from apps.medical.models import MedicalDocument, MedicalDocumentVersion
from apps.outbox.models import OutboxEvent, OutboxEventType, OutboxStatus
from apps.reception.admin import (
    DailyQueueAdmin,
    _admin_resolve_dailyqueue_clinic_site_id,
)
from apps.reception.models import (
    ClinicSite,
    ConsultingRoom,
    DailyQueue,
    ImportStatus,
    Patient,
    PatientImportBatch,
    PatientImportError,
    PatientFormSession,
    QueueEntry,
    QueueStatus,
    TabletDevice,
)
from apps.reception.process_types import ProcessType
from apps.core.api_utils import assign_group_to_test_user
from apps.core.exceptions import DomainError
from apps.reception.services import (
    create_or_update_patient_manual,
    create_queue_entry,
    get_or_create_tablet_device_by_android_id,
    issue_tablet_session_latest_wins,
)
from apps.operations.prom_metrics import build_metrics_payload
from apps.reception.phone_utils import (
    normalize_phone,
    normalize_phone_for_patient_storage,
)
from apps.reception.xlsx_import import (
    XlsxImportErrorCode,
    _audit_xlsx_import_finished,
    _cleanup_clinic_name,
    _parse_date,
    _split_full_name,
    _title_case_name,
    find_patient_for_import,
    process_patient_xlsx_import_batch,
)
from apps.users.models import StaffUser


def _purge_seed_clinic_data() -> None:
    mod = importlib.import_module(
        "apps.reception.migrations.0030_purge_seed_clinics_demo_muc"
    )
    mod.purge_seed_clinic_data(django_apps)


class ReceptionServicesTests(TestCase):
    def setUp(self) -> None:
        self.reception_user = StaffUser.objects.create_user(
            username="reception",
            email="reception@example.com",
            password="safe-password",
            is_staff=True,
        )
        assign_group_to_test_user(self.reception_user, "Reception")
        self.clinic = ClinicSite.objects.create(code="BER", name="Berlin")
        self.room = ConsultingRoom.objects.create(
            clinic_site=self.clinic,
            code="R1",
            name="Room 1",
        )
        self.daily_queue = DailyQueue.objects.create(
            queue_date=timezone.now().date(),
            clinic_site=self.clinic,
            consulting_room=self.room,
            status=QueueStatus.OPEN,
            created_by_user=self.reception_user,
        )

    def test_patient_save_normalizes_name_casing(self) -> None:
        patient = Patient.objects.create(
            first_name="aLeXanDra",
            last_name="nIzhENKO",
            date_of_birth=date(1991, 6, 6),
            phone="+48111222333",
            email="alex@example.com",
        )
        patient.refresh_from_db()
        self.assertEqual(patient.first_name, "Alexandra")
        self.assertEqual(patient.last_name, "Nizhenko")

    def test_find_patient_for_import_matches_title_cased_storage(self) -> None:
        Patient.objects.create(
            first_name="JAN",
            last_name="KOWALSKI",
            date_of_birth=date(1988, 10, 10),
            phone="+48777888910",
            email="jan@example.com",
        )
        found = find_patient_for_import(
            first_name="jan",
            last_name="kowalski",
            phone="+48 777 888 910",
            date_of_birth=date(1988, 10, 10),
        )
        self.assertIsNotNone(found)
        assert found is not None  # narrow for mypy
        self.assertEqual(found.first_name, "Jan")
        self.assertEqual(found.last_name, "Kowalski")

    def test_patient_save_preserves_anonymized_name_sentinel(self) -> None:
        patient = Patient.objects.create(
            first_name="ANONYMIZED",
            last_name="ANONYMIZED",
            date_of_birth=None,
            phone="49999999999",
            email="anon@example.com",
        )
        Patient.objects.filter(pk=patient.pk).update(anonymized_at=timezone.now())
        patient.refresh_from_db()
        patient.save()
        patient.refresh_from_db()
        self.assertEqual(patient.first_name, "ANONYMIZED")
        self.assertEqual(patient.last_name, "ANONYMIZED")

    def test_patient_save_skips_name_normalize_when_last_name_anonymized_sentinel(
        self,
    ) -> None:
        patient = Patient.objects.create(
            first_name="Real",
            last_name="ANONYMIZED",
            date_of_birth=date(1988, 1, 1),
            phone="48111222350",
            email="partial@example.com",
        )
        patient.phone = "48111222351"
        patient.save(update_fields=["phone"])
        patient.refresh_from_db()
        self.assertEqual(patient.first_name, "Real")
        self.assertEqual(patient.last_name, "ANONYMIZED")

    def test_create_or_update_patient_manual_allows_missing_doctolib_id(self) -> None:
        patient = create_or_update_patient_manual(
            first_name="Jan",
            last_name="Nowak",
            date_of_birth=date(1990, 1, 1),
            phone="+48123123123",
            email="jan.nowak@example.com",
            doctolib_patient_id=None,
            created_or_updated_by_user_id=self.reception_user.id,
        )

        self.assertIsNone(patient.doctolib_patient_id)
        self.assertEqual(patient.first_name, "Jan")
        self.assertEqual(patient.phone, "48123123123")

    def test_create_or_update_patient_manual_blocks_stale_anonymized_phone(
        self,
    ) -> None:
        stale = Patient.objects.create(
            first_name="ANONYMIZED",
            last_name="ANONYMIZED",
            date_of_birth=date(1970, 1, 1),
            phone="48777888906",
            email="stale@example.com",
        )
        Patient.objects.filter(pk=stale.pk).update(anonymized_at=timezone.now())

        with self.assertRaises(DomainError) as ctx:
            create_or_update_patient_manual(
                first_name="Nowy",
                last_name="Pacjent",
                date_of_birth=date(1990, 2, 2),
                phone="+48 777 888 906",
                email="nowy@example.com",
                created_or_updated_by_user_id=self.reception_user.id,
            )
        self.assertEqual(
            ctx.exception.api_message_key,
            "other.domain.import_patient_anonymized_same_phone",
        )
        self.assertEqual(Patient.objects.count(), 1)

    def test_patient_identity_unique_constraint_blocks_duplicate_patient(self) -> None:
        Patient.objects.create(
            first_name="Anna",
            last_name="Kowalska",
            date_of_birth=date(1985, 5, 5),
            phone="+48999999999",
            email="anna.k@example.com",
        )

        with self.assertRaises(IntegrityError):
            Patient.objects.create(
                first_name="Anna",
                last_name="Kowalska",
                date_of_birth=date(1985, 5, 5),
                phone="+48999999999",
                email="other@example.com",
            )

    def test_two_patients_same_phone_different_identity_allowed(self) -> None:
        shared_phone = "+48988887777"
        stored_phone = normalize_phone_for_patient_storage(shared_phone)
        Patient.objects.create(
            first_name="Hans",
            last_name="Müller",
            date_of_birth=date(1970, 3, 12),
            phone=shared_phone,
            email="hans@example.com",
        )
        son = Patient.objects.create(
            first_name="Peter",
            last_name="Müller",
            date_of_birth=date(2000, 7, 21),
            phone=shared_phone,
            email="peter@example.com",
        )
        self.assertEqual(Patient.objects.filter(phone=stored_phone).count(), 2)
        self.assertEqual(son.first_name, "Peter")

    def test_find_active_patients_by_phone_excludes_inactive(self) -> None:
        from apps.reception.patient_identity import find_active_patients_by_phone

        phone = normalize_phone_for_patient_storage("+491709998877")
        active = Patient.objects.create(
            first_name="Active",
            last_name="User",
            date_of_birth=date(1980, 1, 1),
            phone=phone,
            email="active@example.com",
            is_active=True,
        )
        Patient.objects.create(
            first_name="Inactive",
            last_name="User",
            date_of_birth=date(1981, 2, 2),
            phone=phone,
            email="inactive@example.com",
            is_active=False,
        )
        found = find_active_patients_by_phone(phone)
        self.assertEqual([p.id for p in found], [active.id])

    def test_shared_phone_warning_uses_same_lookup_variants_as_portal(self) -> None:
        from apps.reception.patient_identity import (
            build_shared_phone_warnings,
            find_active_patients_by_phone,
        )

        stored = normalize_phone_for_patient_storage("+491701112233")
        existing = Patient.objects.create(
            first_name="Anna",
            last_name="Kowalska",
            date_of_birth=date(1975, 1, 1),
            phone=stored,
            email="anna.shared@example.com",
        )
        others = find_active_patients_by_phone("+49 170 111 2233")
        self.assertEqual(len(others), 1)
        self.assertEqual(others[0].id, existing.id)
        warnings = build_shared_phone_warnings(
            phone="+49 170-111 2233",
            exclude_patient_id=uuid.uuid4(),
        )
        self.assertEqual(len(warnings), 1)
        self.assertEqual(warnings[0]["other_patients"][0]["id"], str(existing.id))

    def test_doctolib_patient_id_remains_unique(self) -> None:
        Patient.objects.create(
            first_name="Anna",
            last_name="Kowalska",
            date_of_birth=date(1985, 5, 5),
            phone="+48999999999",
            email="anna.k@example.com",
            doctolib_patient_id="DOC-123",
        )

        with self.assertRaises(IntegrityError):
            Patient.objects.create(
                first_name="Other",
                last_name="Patient",
                date_of_birth=date(1990, 1, 1),
                phone="+48111111111",
                email="other@example.com",
                doctolib_patient_id="DOC-123",
            )

    def test_create_queue_entry_auto_assigns_next_position(self) -> None:
        patient_one = Patient.objects.create(
            first_name="P1",
            last_name="Test",
            date_of_birth=date(1991, 1, 1),
            phone="+48111111111",
            email="p1@example.com",
            doctolib_patient_id="DOC-P1",
        )
        patient_two = Patient.objects.create(
            first_name="P2",
            last_name="Test",
            date_of_birth=date(1992, 2, 2),
            phone="+48222222222",
            email="p2@example.com",
            doctolib_patient_id="DOC-P2",
        )

        first = create_queue_entry(
            daily_queue_id=self.daily_queue.id,
            patient_id=patient_one.id,
            created_by_user_id=self.reception_user.id,
        )
        second = create_queue_entry(
            daily_queue_id=self.daily_queue.id,
            patient_id=patient_two.id,
            created_by_user_id=self.reception_user.id,
        )

        self.assertEqual(first.position_no, 1)
        self.assertEqual(second.position_no, 2)

    def test_issue_tablet_session_latest_wins_switches_active_session(self) -> None:
        patient = Patient.objects.create(
            first_name="Tablet",
            last_name="Patient",
            date_of_birth=date(1993, 3, 3),
            phone="+48333333333",
            email="tablet@example.com",
            doctolib_patient_id="DOC-P3",
        )
        queue_entry = create_queue_entry(
            daily_queue_id=self.daily_queue.id,
            patient_id=patient.id,
            created_by_user_id=self.reception_user.id,
        )

        first_result = issue_tablet_session_latest_wins(
            queue_entry_id=queue_entry.id,
            created_by_user_id=self.reception_user.id,
            form_locale="de-DE",
        )
        second_result = issue_tablet_session_latest_wins(
            queue_entry_id=queue_entry.id,
            created_by_user_id=self.reception_user.id,
            form_locale="en-GB",
        )

        queue_entry.refresh_from_db()
        self.assertEqual(queue_entry.active_session_id, second_result.session_id)
        self.assertNotEqual(first_result.session_id, second_result.session_id)

        self.assertEqual(
            PatientFormSession.objects.filter(queue_entry=queue_entry).count(),
            2,
        )
        self.assertEqual(first_result.intake_form_id, second_result.intake_form_id)
        intake_form = PatientIntakeForm.objects.get(queue_entry=queue_entry)
        self.assertEqual(intake_form.session_id, second_result.session_id)

    def test_patients_api_view_doctor_filtered(self) -> None:
        doctor_user = StaffUser.objects.create_user(
            username="doc_test",
            email="doc_test@example.com",
            password="pwd",
            is_staff=True,
        )
        assign_group_to_test_user(doctor_user, "Doctor")
        # Assign clinic to doctor
        doctor_user.clinic_sites.add(self.clinic)

        # Create a patient assigned to self.clinic
        patient1 = Patient.objects.create(
            first_name="Test1",
            last_name="Test1",
            date_of_birth=date(1991, 1, 1),
            phone="+48111111111",
            email="test1@example.com",
        )
        patient1.clinic_sites.add(self.clinic)

        # Create a patient NOT assigned to self.clinic
        other_clinic = ClinicSite.objects.create(code="OTH", name="Other")
        patient2 = Patient.objects.create(
            first_name="Test2",
            last_name="Test2",
            date_of_birth=date(1991, 1, 1),
            phone="+48111111112",
            email="test2@example.com",
        )
        patient2.clinic_sites.add(other_clinic)

        client = Client()
        client.force_login(doctor_user)
        response = client.get("/api/v1/patients")
        self.assertEqual(response.status_code, 200)

        items = response.json()["items"]
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["id"], str(patient1.id))


class DailyQueueAdminImportTests(TestCase):
    """Tests for admin import-from-file UI (XLSX upload)."""

    def setUp(self) -> None:
        self.client = Client()
        self.admin_user = StaffUser.objects.create_superuser(
            username="admin-import",
            email="admin-import@example.com",
            password="safe-password",
        )
        self.client.force_login(self.admin_user)

    def test_daily_queue_changelist_contains_import_button(self) -> None:
        response = self.client.get(reverse("admin:reception_dailyqueue_changelist"))

        self.assertEqual(response.status_code, 200)
        import_url = reverse("admin:reception_dailyqueue_import_xlsx")
        self.assertContains(response, import_url)

    def test_import_xlsx_admin_view_renders_form(self) -> None:
        response = self.client.get(reverse("admin:reception_dailyqueue_import_xlsx"))

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn('name="file"', content)
        self.assertIn("multipart/form-data", content)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", content
        )


class DailyQueueAdminDoctorFilterTests(TestCase):
    def test_dailyqueue_admin_autocomplete_excludes_consulting_room(self) -> None:
        admin_obj = DailyQueueAdmin(DailyQueue, AdminSite())
        self.assertEqual(
            admin_obj.autocomplete_fields,
            ("clinic_site", "assigned_doctor"),
        )

    def test_assigned_doctor_field_has_doctor_limit_choices_to(self) -> None:
        field = DailyQueue._meta.get_field("assigned_doctor")
        self.assertEqual(field.get_limit_choices_to(), Q(groups__name="Doctor"))

    def test_assigned_doctor_field_shows_only_doctors(self) -> None:
        doctor = StaffUser.objects.create_user(
            username="doctor-filter",
            email="doctor-filter@example.com",
            password="safe-password",
            is_staff=True,
        )
        receptionist = StaffUser.objects.create_user(
            username="reception-filter",
            email="reception-filter@example.com",
            password="safe-password",
            is_staff=True,
        )
        assign_group_to_test_user(doctor, "Doctor")
        assign_group_to_test_user(receptionist, "Reception")

        admin_obj = DailyQueueAdmin(DailyQueue, AdminSite())
        request = RequestFactory().get("/admin/reception/dailyqueue/add/")
        request.user = doctor
        db_field = DailyQueue._meta.get_field("assigned_doctor")

        formfield = admin_obj.formfield_for_foreignkey(db_field, request)
        user_ids = set(formfield.queryset.values_list("id", flat=True))

        self.assertIn(doctor.id, user_ids)
        self.assertNotIn(receptionist.id, user_ids)

    def test_daily_queue_clean_requires_consulting_room_same_clinic_site(self) -> None:
        site_a = ClinicSite.objects.create(code="A", name="Site A")
        site_b = ClinicSite.objects.create(code="B", name="Site B")
        room_b = ConsultingRoom.objects.create(
            clinic_site=site_b, code="RB", name="Room B"
        )
        user = StaffUser.objects.create_user(
            username="dq-clean",
            email="dq-clean@example.com",
            password="safe-password",
            is_staff=True,
        )
        dq = DailyQueue(
            queue_date=date(2026, 4, 1),
            clinic_site=site_a,
            consulting_room=room_b,
            created_by_user=user,
        )
        with self.assertRaises(ValidationError) as ctx:
            dq.full_clean()
        self.assertIn("consulting_room", ctx.exception.error_dict)

    def test_admin_resolve_clinic_site_prefers_post_over_obj(self) -> None:
        site_a = ClinicSite.objects.create(code="PA", name="Post A")
        site_b = ClinicSite.objects.create(code="PB", name="Post B")
        user = StaffUser.objects.create_user(
            username="dq-post",
            email="dq-post@example.com",
            password="safe-password",
            is_staff=True,
        )
        dq = DailyQueue.objects.create(
            queue_date=date(2026, 4, 2),
            clinic_site=site_b,
            consulting_room=ConsultingRoom.objects.create(
                clinic_site=site_b, code="RPB", name="R B"
            ),
            created_by_user=user,
        )
        request = RequestFactory().post(
            "/admin/reception/dailyqueue/change/",
            {"clinic_site": str(site_a.id)},
        )
        self.assertEqual(
            _admin_resolve_dailyqueue_clinic_site_id(request, dq), site_a.id
        )

        get_req = RequestFactory().get("/admin/reception/dailyqueue/change/")
        self.assertEqual(
            _admin_resolve_dailyqueue_clinic_site_id(get_req, dq), site_b.id
        )


class XlsxImportParsingTests(TestCase):
    def test_cleanup_clinic_name_removes_trailing_weekday_and_date(self) -> None:
        cleaned = _cleanup_clinic_name("Kreutzigerstraße Freitag, 6. März")
        self.assertEqual(cleaned, "Kreutzigerstraße")

    def test_parse_date_accepts_dob_with_age_suffix(self) -> None:
        parsed = _parse_date("4.07.1996 (30 Jahre)")
        self.assertEqual(parsed, date(1996, 7, 4))

    def test_split_full_name_removes_title_and_symbol(self) -> None:
        first_name, last_name = _split_full_name("Herr FRITSCHE Sebastian @")
        self.assertEqual(first_name, "Sebastian")
        self.assertEqual(last_name, "FRITSCHE")

    def test_split_full_name_handles_frau_format(self) -> None:
        first_name, last_name = _split_full_name("Frau JURGA Jolina")
        self.assertEqual(first_name, "Jolina")
        self.assertEqual(last_name, "JURGA")

    def test_title_case_name_normalizes_case(self) -> None:
        self.assertEqual(_title_case_name("aLeXanDra"), "Alexandra")
        self.assertEqual(_title_case_name("nIzhENKO"), "Nizhenko")
        self.assertEqual(_title_case_name("o'NEIL-smITH"), "O'Neil-Smith")


class TabletWebLoginLastSeenTests(TestCase):
    def test_tablet_login_with_android_id_sets_last_seen_at(self) -> None:
        client = Client()
        user = StaffUser.objects.create_user(
            username="tablet-login-seen",
            email="tablet-login-seen@example.com",
            password="safe-password",
            is_staff=True,
        )
        assign_group_to_test_user(user, "Tablet")
        device = TabletDevice.objects.create(
            android_id="web-login-android-seen", is_active=True
        )
        self.assertIsNone(device.last_seen_at)
        response = client.post(
            "/tablet/login/",
            data={
                "username": "tablet-login-seen",
                "password": "safe-password",
                "android_id": "web-login-android-seen",
            },
        )
        self.assertEqual(response.status_code, 302)
        device.refresh_from_db()
        self.assertIsNotNone(device.last_seen_at)


class TabletDeviceAutoRegistrationTests(TestCase):
    def test_auto_registered_tablet_device_has_no_default_clinic_site(self) -> None:
        device, created = get_or_create_tablet_device_by_android_id(
            android_id="auto-reg-no-clinic"
        )
        self.assertTrue(created)
        self.assertIsNone(device.clinic_site_id)


class TabletHomeClinicScopeTests(TestCase):
    def setUp(self) -> None:
        self.client = Client()
        self.tablet_user = StaffUser.objects.create_user(
            username="tablet-home-scope",
            email="tablet-home-scope@example.com",
            password="safe-password",
            is_staff=True,
        )
        assign_group_to_test_user(self.tablet_user, "Tablet")
        self.clinic_a = ClinicSite.objects.create(code="TH-A", name="Tablet Home A")
        self.clinic_b = ClinicSite.objects.create(code="TH-B", name="Tablet Home B")
        self.room_a = ConsultingRoom.objects.create(
            clinic_site=self.clinic_a, code="THA-1", name="A1"
        )
        self.room_b = ConsultingRoom.objects.create(
            clinic_site=self.clinic_b, code="THB-1", name="B1"
        )
        DailyQueue.objects.create(
            queue_date=timezone.localdate(),
            clinic_site=self.clinic_a,
            consulting_room=self.room_a,
            created_by_user=self.tablet_user,
        )
        DailyQueue.objects.create(
            queue_date=timezone.localdate(),
            clinic_site=self.clinic_b,
            consulting_room=self.room_b,
            created_by_user=self.tablet_user,
        )
        self.tablet_user.clinic_sites.add(self.clinic_a)

    def test_tablet_home_without_device_session_is_scoped_to_assigned_clinics(
        self,
    ) -> None:
        self.client.force_login(self.tablet_user)
        response = self.client.get("/tablet/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Tablet Home A")
        self.assertNotContains(response, "Tablet Home B")


class ReceptionDashboardScopeTests(TestCase):
    def setUp(self) -> None:
        self.client = Client()
        self.admin_user = StaffUser.objects.create_user(
            username="dashboard-admin",
            email="dashboard-admin@example.com",
            password="safe-password",
            is_staff=True,
        )
        assign_group_to_test_user(self.admin_user, "Admin")
        self.reception_user = StaffUser.objects.create_user(
            username="dashboard-reception",
            email="dashboard-reception@example.com",
            password="safe-password",
            is_staff=True,
        )
        assign_group_to_test_user(self.reception_user, "Reception")
        self.clinic_a = ClinicSite.objects.create(
            code="RDSA", name="Reception Dashboard A"
        )
        self.clinic_b = ClinicSite.objects.create(
            code="RDSB", name="Reception Dashboard B"
        )
        self.room_a = ConsultingRoom.objects.create(
            clinic_site=self.clinic_a, code="RA", name="Room A"
        )
        self.room_b = ConsultingRoom.objects.create(
            clinic_site=self.clinic_b, code="RB", name="Room B"
        )
        self.reception_user.clinic_sites.add(self.clinic_a)

    def _create_failed_outbox_event(
        self, *, clinic: ClinicSite, room: ConsultingRoom, suffix: str
    ) -> OutboxEvent:
        patient = Patient.objects.create(
            first_name=f"Dash{suffix}",
            last_name="Patient",
            date_of_birth=date(1990, 1, 1),
            phone=f"49999{suffix}",
            email=f"dash-{suffix}@example.com",
            doctolib_patient_id=f"DASH-{suffix}",
        )
        queue = DailyQueue.objects.create(
            queue_date=timezone.now().date(),
            clinic_site=clinic,
            consulting_room=room,
            created_by_user=self.admin_user,
        )
        entry = QueueEntry.objects.create(
            daily_queue=queue,
            patient=patient,
            position_no=1,
            created_by_user=self.admin_user,
        )
        session = PatientFormSession.objects.create(
            queue_entry=entry,
            form_locale="de-DE",
            expires_at=timezone.now() + timedelta(minutes=60),
            created_by_user=self.admin_user,
        )
        intake_form = PatientIntakeForm.objects.create(
            queue_entry=entry,
            session=session,
            form_status=IntakeStatus.IN_PROGRESS,
        )
        medical_document = MedicalDocument.objects.create(
            queue_entry=entry,
            intake_form=intake_form,
            created_by_user=self.admin_user,
        )
        version = MedicalDocumentVersion.objects.create(
            medical_document=medical_document,
            version_no=1,
            medical_payload={"schema_version": 1},
        )
        return OutboxEvent.objects.create(
            medical_document_version=version,
            aggregate_id=version.id,
            event_type=OutboxEventType.GENERATE_PDF,
            payload={"schema_version": 1},
            status=OutboxStatus.FAILED,
            error_message="failed",
        )

    def test_reception_dashboard_outbox_errors_are_scoped_to_assigned_clinics(
        self,
    ) -> None:
        event_in_scope = self._create_failed_outbox_event(
            clinic=self.clinic_a, room=self.room_a, suffix="11"
        )
        event_out_of_scope = self._create_failed_outbox_event(
            clinic=self.clinic_b, room=self.room_b, suffix="22"
        )
        self.client.force_login(self.reception_user)
        response = self.client.get(reverse("admin_reception_dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response, str(event_in_scope.medical_document_version.medical_document_id)
        )
        self.assertNotContains(
            response,
            str(event_out_of_scope.medical_document_version.medical_document_id),
        )


def _write_xlsx_without_valid_patient_header(
    path: Path,
    *,
    queue_date: date,
    standort_name: str,
) -> None:
    """Date + Standort present, but no row with phone/email/name headers."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = queue_date.strftime("%d.%m.%Y")
    ws["A2"] = f"Standort: {standort_name}"
    ws.cell(4, 1, "not_a_patient_header_row")
    wb.save(path)
    wb.close()


def _write_minimal_patient_xlsx(
    path: Path,
    *,
    queue_date: date,
    standort_name: str,
    data_rows: list[tuple[str, ...]],
    extra_header: str | None = None,
) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = queue_date.strftime("%d.%m.%Y")
    ws["A2"] = f"Standort: {standort_name}"
    header_row = 4
    headers = ["first_name", "last_name", "date_of_birth", "phone", "email"]
    if extra_header:
        headers.append(extra_header)
    for col, title in enumerate(headers, start=1):
        ws.cell(header_row, col, title)
    r = header_row + 1
    for data in data_rows:
        for col, val in enumerate(data, start=1):
            ws.cell(r, col, val)
        r += 1
    wb.save(path)
    wb.close()


class PatientXlsxImportTests(TestCase):
    """Integration tests for process_patient_xlsx_import_batch (import plan)."""

    def setUp(self) -> None:
        _purge_seed_clinic_data()
        self.user = StaffUser.objects.create_user(
            username="xlsx-import-tester",
            email="xlsx-import@example.com",
            password="safe-password",
            is_staff=True,
        )
        self.clinic = ClinicSite.objects.create(
            code="XIIMP", name="Xlsx Import Test Clinic München"
        )
        self.room = ConsultingRoom.objects.create(
            clinic_site=self.clinic, code="XI1", name="Import room"
        )
        self.clinic.pdf_import_default_consulting_room = self.room
        self.clinic.save(update_fields=["pdf_import_default_consulting_room"])
        self.import_day = date(2026, 6, 10)

    def _run_import(
        self,
        data_rows: list[tuple[str, ...]],
        *,
        extra_header: str | None = None,
    ) -> PatientImportBatch:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = Path(tmp.name)
        try:
            _write_minimal_patient_xlsx(
                path,
                queue_date=self.import_day,
                standort_name=self.clinic.name,
                data_rows=data_rows,
                extra_header=extra_header,
            )
            batch = PatientImportBatch.objects.create(
                source_file_name="rows.xlsx",
                source_file_sha256="a" * 64,
                created_by_user=self.user,
            )
            process_patient_xlsx_import_batch(
                batch_id=batch.id, stored_file_path=str(path)
            )
            batch.refresh_from_db()
            return batch
        finally:
            path.unlink(missing_ok=True)

    def test_validate_patient_names_for_import_rejects_placeholder(self) -> None:
        from apps.core.exceptions import DomainError
        from apps.reception.patient_identity import validate_patient_names_for_import

        with self.assertRaises(DomainError) as ctx:
            validate_patient_names_for_import(first_name="—", last_name="Kowalski")
        self.assertEqual(
            ctx.exception.api_message_key,
            "other.domain.import_missing_patient_name",
        )

    def test_patient_is_active_record_false_for_anonymized_at(self) -> None:
        from apps.reception.patient_identity import (
            patient_is_active_record,
            patient_is_import_anonymized,
        )

        patient = Patient.objects.create(
            first_name="Ewa",
            last_name="Test",
            date_of_birth=date(1980, 1, 1),
            phone="48111222352",
            email="ewa.anon@example.com",
        )
        Patient.objects.filter(pk=patient.pk).update(anonymized_at=timezone.now())
        patient.refresh_from_db()
        self.assertTrue(patient_is_import_anonymized(patient))
        self.assertFalse(patient_is_active_record(patient))

    def test_find_patient_for_import_none_for_inactive(self) -> None:
        Patient.objects.create(
            first_name="Inactive",
            last_name="Import",
            date_of_birth=date(1980, 2, 2),
            phone="48111222353",
            email="inactive.import@example.com",
            is_active=False,
        )
        self.assertIsNone(
            find_patient_for_import(
                first_name="Inactive",
                last_name="Import",
                phone="48111222353",
                date_of_birth=date(1980, 2, 2),
            )
        )

    def test_iter_patients_matching_phone_dedupes_variant_hits(self) -> None:
        from apps.reception.patient_identity import _iter_patients_matching_phone

        stored = normalize_phone_for_patient_storage("+491701112244")
        Patient.objects.create(
            first_name="Legacy",
            last_name="Plus",
            date_of_birth=date(1975, 3, 3),
            phone=f"+{stored}",
            email="legacy.plus@example.com",
        )
        found = _iter_patients_matching_phone("+49 170 111 2244")
        self.assertEqual(len(found), 1)

    def test_assert_phone_not_blocked_when_exclude_patient_already_has_phone(
        self,
    ) -> None:
        from apps.reception.patient_identity import (
            assert_phone_not_blocked_by_stale_anonymized,
        )

        phone = normalize_phone_for_patient_storage("+48 777 888 907")
        stale = Patient.objects.create(
            first_name="ANONYMIZED",
            last_name="ANONYMIZED",
            date_of_birth=date(1970, 1, 1),
            phone=phone,
            email="stale.exclude@example.com",
        )
        Patient.objects.filter(pk=stale.pk).update(anonymized_at=timezone.now())
        active = Patient.objects.create(
            first_name="Active",
            last_name="Holder",
            date_of_birth=date(1990, 1, 1),
            phone=phone,
            email="active.holder@example.com",
        )
        assert_phone_not_blocked_by_stale_anonymized(
            phone=phone,
            exclude_patient_id=active.id,
        )

    def test_iter_patients_matching_phone_empty_returns_none(self) -> None:
        from apps.reception.patient_identity import _iter_patients_matching_phone

        self.assertEqual(_iter_patients_matching_phone(""), [])

    def test_find_patient_for_import_none_for_anonymized(self) -> None:
        p = Patient.objects.create(
            first_name="ANONYMIZED",
            last_name="ANONYMIZED",
            date_of_birth=date(1980, 1, 1),
            phone="48111222301",
            email="anon@example.com",
        )
        Patient.objects.filter(pk=p.pk).update(anonymized_at=timezone.now())
        p.refresh_from_db()
        self.assertIsNone(
            find_patient_for_import(
                first_name="ANONYMIZED",
                last_name="ANONYMIZED",
                phone="48111222301",
                date_of_birth=date(1980, 1, 1),
            )
        )

    def test_find_patient_for_import_returns_active(self) -> None:
        p = Patient.objects.create(
            first_name="Ewa",
            last_name="K",
            date_of_birth=date(1980, 1, 1),
            phone="48111222302",
            email="ewa@example.com",
        )
        found = find_patient_for_import(
            first_name="Ewa",
            last_name="K",
            phone="48111222302",
            date_of_birth=date(1980, 1, 1),
        )
        self.assertIsNotNone(found)
        assert found is not None
        self.assertEqual(found.id, p.id)

    def test_import_new_patient(self) -> None:
        batch = self._run_import(
            [("Nina", "Nowa", "01.01.1992", "+48 777 888 901", "nina901@example.com")],
        )
        self.assertEqual(batch.status, ImportStatus.COMPLETED)
        self.assertEqual(batch.inserted_rows, 1)
        self.assertEqual(batch.matched_rows, 0)
        self.assertEqual(batch.error_rows, 0)
        norm = normalize_phone_for_patient_storage("+48 777 888 901")
        self.assertEqual(Patient.objects.filter(phone=norm).count(), 1)
        payload = build_metrics_payload()
        self.assertIn(b"cogitomedica_import_batches_total", payload)

    @patch(
        "apps.reception.xlsx_import.create_or_update_patient_manual",
        side_effect=RuntimeError("import boom"),
    )
    def test_import_patient_create_exception_records_row_error(
        self, _mock_create
    ) -> None:
        batch = self._run_import(
            [("Fail", "Row", "01.01.1990", "+48 777 888 912", "fail@example.com")],
        )
        self.assertEqual(batch.status, ImportStatus.COMPLETED_WITH_ERRORS)
        self.assertEqual(batch.error_rows, 1)
        self.assertEqual(Patient.objects.count(), 0)
        err = PatientImportError.objects.get(batch=batch)
        self.assertEqual(err.error_code, XlsxImportErrorCode.INVALID_ROW_FORMAT)
        self.assertIn("import boom", err.error_message)

    def test_import_existing_patient_reuses_record_same_identity(self) -> None:
        Patient.objects.create(
            first_name="Stary",
            last_name="Pacjent",
            date_of_birth=date(1990, 1, 1),
            phone="48777888902",
            email="stary@example.com",
        )
        batch = self._run_import(
            [
                (
                    "Stary",
                    "Pacjent",
                    "01.01.1990",
                    "+48 777 888 902",
                    "nowyemail@example.com",
                )
            ],
        )
        self.assertEqual(batch.status, ImportStatus.COMPLETED)
        self.assertEqual(Patient.objects.count(), 1)
        self.assertEqual(batch.inserted_rows, 0)
        self.assertEqual(batch.matched_rows, 1)
        p = Patient.objects.get()
        self.assertEqual(p.first_name, "Stary")
        self.assertEqual(QueueEntry.objects.filter(patient=p).count(), 1)

    def test_import_same_phone_different_identity_creates_new_patient(self) -> None:
        Patient.objects.create(
            first_name="Stary",
            last_name="Pacjent",
            date_of_birth=date(1990, 1, 1),
            phone="48777888902",
            email="stary@example.com",
        )
        batch = self._run_import(
            [
                (
                    "Inny",
                    "Import",
                    "15.05.1995",
                    "+48 777 888 902",
                    "nowyemail@example.com",
                )
            ],
        )
        self.assertEqual(batch.status, ImportStatus.COMPLETED)
        self.assertEqual(Patient.objects.count(), 2)
        self.assertEqual(batch.inserted_rows, 1)
        self.assertEqual(batch.matched_rows, 0)

    def test_import_anonymized_patient_creates_new(self) -> None:
        old = Patient.objects.create(
            first_name="Jan",
            last_name="Doe",
            date_of_birth=date(1985, 3, 3),
            phone="48777888903",
            email="jan@example.com",
        )
        sentinel = str(old.id.int)[:20]
        Patient.objects.filter(pk=old.pk).update(
            phone=sentinel,
            first_name="ANONYMIZED",
            last_name="ANONYMIZED",
            date_of_birth=None,
            anonymized_at=timezone.now(),
        )
        batch = self._run_import(
            [
                (
                    "Powrot",
                    "Pacjent",
                    "10.10.1988",
                    "+48 777 888 903",
                    "powrot@example.com",
                )
            ],
        )
        self.assertEqual(batch.status, ImportStatus.COMPLETED)
        self.assertEqual(batch.inserted_rows, 1)
        self.assertEqual(batch.matched_rows, 0)
        self.assertEqual(Patient.objects.count(), 2)
        self.assertTrue(
            Patient.objects.filter(phone=normalize_phone("+48 777 888 903")).exists()
        )

    def test_import_duplicate_in_file(self) -> None:
        row = ("A", "B", "01.01.1991", "+48 777 888 904", "dup@example.com")
        batch = self._run_import([row, row])
        self.assertEqual(batch.status, ImportStatus.COMPLETED_WITH_ERRORS)
        self.assertEqual(batch.error_rows, 1)
        self.assertEqual(batch.inserted_rows, 1)
        self.assertEqual(batch.matched_rows, 0)
        PatientImportError.objects.get(
            batch=batch,
            error_code=XlsxImportErrorCode.DUPLICATE_IN_FILE,
        )

    def test_import_same_phone_three_different_identities_in_one_file(self) -> None:
        """Doctolib-style rows: shared family phone, different names/DOBs — no DUPLICATE_IN_FILE."""
        batch = self._run_import(
            [
                ("A", "B", "4.07.1996 (30 Jahre)", "664412709", "email@gmail.com"),
                ("B", "B", "5.07.1996 (30 Jahre)", "664412709", "email@gmail.com"),
                ("C", "B", "6.07.1996 (30 Jahre)", "664412709", "email@gmail.com"),
            ],
        )
        self.assertEqual(batch.status, ImportStatus.COMPLETED)
        self.assertEqual(batch.error_rows, 0)
        self.assertEqual(batch.inserted_rows, 3)
        self.assertEqual(Patient.objects.count(), 3)
        self.assertFalse(
            PatientImportError.objects.filter(
                batch=batch,
                error_code=XlsxImportErrorCode.DUPLICATE_IN_FILE,
            ).exists()
        )

    def test_reimport_same_daily_queue_skips_existing_queue_entry(self) -> None:
        row = (
            "Erika",
            "Mustermann",
            "01.01.1991",
            "+48 777 888 906",
            "erika@example.com",
        )

        first_batch = self._run_import([row])
        second_batch = self._run_import([row])

        self.assertEqual(first_batch.status, ImportStatus.COMPLETED)
        self.assertEqual(first_batch.inserted_rows, 1)
        self.assertEqual(first_batch.skipped_already_present_count, 0)

        self.assertEqual(second_batch.status, ImportStatus.COMPLETED)
        self.assertEqual(second_batch.inserted_rows, 0)
        self.assertEqual(second_batch.matched_rows, 0)
        self.assertEqual(second_batch.error_rows, 0)
        self.assertEqual(second_batch.skipped_already_present_count, 1)
        self.assertEqual(Patient.objects.count(), 1)
        self.assertEqual(QueueEntry.objects.count(), 1)

    def test_import_fails_when_no_valid_patient_header_row(self) -> None:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = Path(tmp.name)
        try:
            _write_xlsx_without_valid_patient_header(
                path,
                queue_date=self.import_day,
                standort_name=self.clinic.name,
            )
            batch = PatientImportBatch.objects.create(
                source_file_name="bad-header.xlsx",
                source_file_sha256="d" * 64,
                created_by_user=self.user,
            )
            process_patient_xlsx_import_batch(
                batch_id=batch.id, stored_file_path=str(path)
            )
            batch.refresh_from_db()
            self.assertEqual(batch.status, ImportStatus.FAILED)
        finally:
            path.unlink(missing_ok=True)

    def test_import_domain_error_finishes_batch(self) -> None:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = Path(tmp.name)
        try:
            _write_minimal_patient_xlsx(
                path,
                queue_date=self.import_day,
                standort_name=self.clinic.name,
                data_rows=[],
            )
            batch = PatientImportBatch.objects.create(
                source_file_name="domain.xlsx",
                source_file_sha256="e" * 64,
                created_by_user=self.user,
            )
            with patch(
                "apps.reception.xlsx_import._extract_file_metadata",
                side_effect=DomainError("forced domain error"),
            ):
                process_patient_xlsx_import_batch(
                    batch_id=batch.id, stored_file_path=str(path)
                )
            batch.refresh_from_db()
            self.assertEqual(batch.status, ImportStatus.FAILED)
        finally:
            path.unlink(missing_ok=True)

    def test_import_openpyxl_unexpected_error_finishes_batch(self) -> None:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = Path(tmp.name)
        try:
            path.write_bytes(b"not an xlsx")
            batch = PatientImportBatch.objects.create(
                source_file_name="broken.xlsx",
                source_file_sha256="f" * 64,
                created_by_user=self.user,
            )
            process_patient_xlsx_import_batch(
                batch_id=batch.id, stored_file_path=str(path)
            )
            batch.refresh_from_db()
            self.assertEqual(batch.status, ImportStatus.FAILED)
        finally:
            path.unlink(missing_ok=True)

    def test_import_rejects_placeholder_or_empty_patient_names(self) -> None:
        from apps.core.domain_messages import domain_message

        missing_name_msg = domain_message("other.domain.import_missing_patient_name")
        cases = (
            (
                ("—", "Nowak", "01.01.1990", "+48 777 888 907", "dash@example.com"),
                XlsxImportErrorCode.INVALID_ROW_FORMAT,
                missing_name_msg,
            ),
            (
                ("", "Pusty", "02.02.1991", "+48 777 888 908", "empty@example.com"),
                XlsxImportErrorCode.MISSING_REQUIRED_FIELD,
                None,
            ),
            (
                ("Jan", "-", "03.03.1992", "+48 777 888 909", "hyphen@example.com"),
                XlsxImportErrorCode.INVALID_ROW_FORMAT,
                missing_name_msg,
            ),
        )
        for row, expected_code, expected_message in cases:
            with self.subTest(row=row):
                batch = self._run_import([row])
                self.assertEqual(batch.status, ImportStatus.COMPLETED_WITH_ERRORS)
                self.assertEqual(batch.inserted_rows, 0)
                self.assertEqual(batch.error_rows, 1)
                err = PatientImportError.objects.get(batch=batch)
                self.assertEqual(err.error_code, expected_code)
                if expected_message is not None:
                    self.assertEqual(err.error_message, expected_message)
                self.assertEqual(Patient.objects.count(), 0)

    def test_import_stale_anonymized_same_phone_errors(self) -> None:
        p = Patient.objects.create(
            first_name="ANONYMIZED",
            last_name="ANONYMIZED",
            date_of_birth=date(1970, 1, 1),
            phone="48777888905",
            email="z@z.com",
        )
        Patient.objects.filter(pk=p.pk).update(anonymized_at=timezone.now())
        batch = self._run_import(
            [("X", "Y", "01.01.1990", "+48 777 888 905", "x@y.com")],
        )
        self.assertEqual(batch.status, ImportStatus.COMPLETED_WITH_ERRORS)
        self.assertEqual(batch.inserted_rows, 0)
        self.assertEqual(batch.matched_rows, 0)
        self.assertEqual(batch.error_rows, 1)
        PatientImportError.objects.get(
            batch=batch,
            error_code=XlsxImportErrorCode.PATIENT_ANONYMIZED_NEW_RECORD,
        )
        self.assertEqual(Patient.objects.count(), 1)

    def test_audit_finished_event_includes_skipped_already_present_count(self) -> None:
        batch = PatientImportBatch.objects.create(
            source_file_name="audit.xlsx",
            source_file_sha256="1" * 64,
            created_by_user=self.user,
        )
        with patch("apps.reception.xlsx_import.create_audit_event") as mocked_audit:
            _audit_xlsx_import_finished(
                batch,
                context_clinic_site_id=self.clinic.id,
                status=ImportStatus.COMPLETED.value,  # type: ignore[attr-defined]
                inserted_rows=2,
                matched_rows=1,
                skipped_already_present_count=3,
                error_rows=0,
            )

        mocked_audit.assert_called_once()
        metadata = mocked_audit.call_args.kwargs["metadata"]
        self.assertEqual(metadata["inserted_rows"], 2)
        self.assertEqual(metadata["matched_rows"], 1)
        self.assertEqual(metadata["skipped_already_present_count"], 3)
        self.assertEqual(metadata["error_rows"], 0)

    def test_import_v1_without_process_column_creates_standard(self) -> None:
        batch = self._run_import(
            [
                (
                    "Erika",
                    "Mustermann",
                    "01.01.1991",
                    "+48 777 888 907",
                    "erika.v1@example.com",
                )
            ]
        )
        self.assertEqual(batch.status, ImportStatus.COMPLETED)
        entry = QueueEntry.objects.get()
        self.assertEqual(entry.process_type, ProcessType.STANDARD)

    def test_import_v1_duplicate_identity_in_file_is_error(self) -> None:
        row = (
            "Erika",
            "Mustermann",
            "01.01.1991",
            "+48 777 888 908",
            "erika.dup@example.com",
        )
        batch = self._run_import([row, row])
        self.assertEqual(batch.error_rows, 1)
        self.assertEqual(batch.inserted_rows, 1)
        self.assertTrue(
            PatientImportError.objects.filter(
                batch=batch,
                error_code=XlsxImportErrorCode.DUPLICATE_IN_FILE,
            ).exists()
        )

    def test_import_v2_standard_then_telederm_creates_two_entries(self) -> None:
        identity = (
            "Erika",
            "Mustermann",
            "01.01.1991",
            "+48 777 888 909",
            "erika.ab@example.com",
        )
        batch = self._run_import(
            [
                (*identity, "STANDARD"),
                (
                    *identity,
                    "Hautarzt-Videosprechstunde mit professioneller Bilddokumentation",
                ),
            ],
            extra_header="Terminart",
        )
        self.assertEqual(batch.status, ImportStatus.COMPLETED)
        self.assertEqual(batch.inserted_rows, 1)
        self.assertEqual(batch.matched_rows, 1)
        types = set(QueueEntry.objects.values_list("process_type", flat=True))
        self.assertEqual(types, {ProcessType.STANDARD, ProcessType.TELEDERM})

    def test_import_v2_reimport_same_type_skips_second_type_creates(self) -> None:
        identity = (
            "Erika",
            "Mustermann",
            "01.01.1991",
            "+48 777 888 910",
            "erika.re@example.com",
        )
        first = self._run_import(
            [(*identity, "STANDARD")],
            extra_header="Terminart",
        )
        second = self._run_import(
            [
                (*identity, "STANDARD"),
                (*identity, "TELEDERM"),
            ],
            extra_header="Terminart",
        )
        self.assertEqual(first.inserted_rows, 1)
        self.assertEqual(second.skipped_already_present_count, 1)
        self.assertEqual(second.inserted_rows, 0)
        self.assertEqual(second.matched_rows, 1)
        self.assertEqual(QueueEntry.objects.count(), 2)

    def test_import_v2_unknown_cell_falls_back_to_standard(self) -> None:
        batch = self._run_import(
            [
                (
                    "Erika",
                    "Mustermann",
                    "01.01.1991",
                    "+48 777 888 911",
                    "erika.fb@example.com",
                    "Unbekannte Leistung",
                )
            ],
            extra_header="Terminart",
        )
        self.assertEqual(batch.status, ImportStatus.COMPLETED)
        entry = QueueEntry.objects.get()
        self.assertEqual(entry.process_type, ProcessType.STANDARD)


class PurgeSeedClinicDataTests(TestCase):
    def setUp(self) -> None:
        self.user = StaffUser.objects.create_user(
            username="purge-tester",
            email="purge-tester@example.com",
            password="safe-password",
            is_staff=True,
        )

    def test_purge_removes_demo_site_and_seed_patients(self) -> None:
        demo = ClinicSite.objects.create(code="DEMO", name="Demo Seed")
        room = ConsultingRoom.objects.create(clinic_site=demo, code="A1", name="R1")
        dq = DailyQueue.objects.create(
            queue_date=date(2026, 1, 1),
            clinic_site=demo,
            consulting_room=room,
            created_by_user=self.user,
        )
        pat = Patient.objects.create(
            first_name="S",
            last_name="P",
            date_of_birth=date(1990, 1, 1),
            phone="111111111",
            email="s@example.com",
            doctolib_patient_id="DTL-2026-9999",
        )
        QueueEntry.objects.create(
            daily_queue=dq,
            patient=pat,
            position_no=1,
            created_by_user=self.user,
        )
        _purge_seed_clinic_data()
        self.assertFalse(ClinicSite.objects.filter(code="DEMO").exists())
        self.assertFalse(
            Patient.objects.filter(doctolib_patient_id="DTL-2026-9999").exists()
        )

    def test_purge_keeps_non_seed_clinic_and_patients(self) -> None:
        real = ClinicSite.objects.create(code="BER", name="Real")
        room = ConsultingRoom.objects.create(clinic_site=real, code="R1", name="R1")
        dq = DailyQueue.objects.create(
            queue_date=date(2026, 1, 2),
            clinic_site=real,
            consulting_room=room,
            created_by_user=self.user,
        )
        pat = Patient.objects.create(
            first_name="R",
            last_name="E",
            date_of_birth=date(1991, 1, 1),
            phone="222222222",
            email="r@example.com",
            doctolib_patient_id="REAL-999",
        )
        QueueEntry.objects.create(
            daily_queue=dq,
            patient=pat,
            position_no=1,
            created_by_user=self.user,
        )
        _purge_seed_clinic_data()
        self.assertTrue(ClinicSite.objects.filter(code="BER").exists())
        self.assertTrue(Patient.objects.filter(doctolib_patient_id="REAL-999").exists())
