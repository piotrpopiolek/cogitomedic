"""
Patient import from XLSX template.

Queue date and clinic site are extracted from the file content
(Doctolib export header: "Standort ...", date string).
"""

from __future__ import annotations

import hashlib
import logging
import re
import unicodedata
import uuid
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path

from django.conf import settings
from django.utils import timezone

from apps.core.domain_messages import domain_message
from apps.core.exceptions import DomainError
from apps.operations.prom_metrics import record_import_batch_finished
from apps.operations.services import create_audit_event
from apps.reception.models import (
    ClinicSite,
    DailyQueue,
    ImportSourceSystem,
    ImportStatus,
    ImportType,
    Patient,
    PatientImportBatch,
    PatientImportError,
    QueueSource,
)
from apps.reception.phone_utils import normalize_phone_for_patient_storage
from apps.reception.patient_identity import (
    find_patient_for_import,
    normalize_email_for_storage,
    patient_identity_key,
    stale_anonymized_patient_blocks_phone,
    validate_patient_names_for_import,
)
from apps.reception.process_types import ProcessType
from apps.reception.services import (
    active_queue_entry_for_process_exists,
    create_daily_queue,
    create_or_update_patient_manual,
    create_queue_entry,
)

logger = logging.getLogger(__name__)


# --- Error codes (aligned with batch/PatientImportError) ---
class XlsxImportFailure(DomainError):
    def __init__(self, error_code: str, message: str) -> None:
        super().__init__(message)
        self.error_code = error_code


class XlsxImportErrorCode:
    TEMPLATE_HEADER_INVALID = "TEMPLATE_HEADER_INVALID"
    MISSING_IMPORT_DATE = "MISSING_IMPORT_DATE"
    MISSING_CLINIC_NAME = "MISSING_CLINIC_NAME"
    UNKNOWN_CLINIC = "UNKNOWN_CLINIC"
    INVALID_ROW_FORMAT = "INVALID_ROW_FORMAT"
    INVALID_DATE_OF_BIRTH = "INVALID_DATE_OF_BIRTH"
    INVALID_PHONE = "INVALID_PHONE"
    MISSING_REQUIRED_FIELD = "MISSING_REQUIRED_FIELD"
    DUPLICATE_VISIT = "DUPLICATE_VISIT"
    DUPLICATE_IN_FILE = "DUPLICATE_IN_FILE"
    PATIENT_ANONYMIZED_NEW_RECORD = "PATIENT_ANONYMIZED_NEW_RECORD"


# Re-export for tests and callers.
__all__ = [
    "find_patient_for_import",
    "map_xlsx_process_type_cell",
    "process_patient_xlsx_import_batch",
]


# --- Header mapping: possible header labels (normalized) -> internal key ---
HEADER_ALIASES = {
    "first_name": ["first_name", "first name", "imie", "imię", "vorname", "prenom"],
    "last_name": ["last_name", "last name", "nazwisko", "nachname", "nom"],
    "full_name": ["pacjent", "patient", "patientin", "name", "patient:in"],
    "date_of_birth": [
        "dob",
        "date_of_birth",
        "date of birth",
        "data urodzenia",
        "geburtsdatum",
        "birth date",
    ],
    "phone": ["phone", "telefon", "tel", "mobile"],
    "email": ["email", "e-mail", "e-mail-adresse", "mail"],
    "appointment_time": ["godzina", "uhrzeit", "time", "appointment_time", "heure"],
    "address": ["adres", "address", "anschrift", "adresse"],
    "postal_code": ["kod pocztowy", "postal_code", "postleitzahl", "plz", "zip"],
    "city": ["miasto", "city", "ort", "stadt", "wohnort", "locality"],
}

# Exact header match only (B4): substring matching would catch unrelated columns.
PROCESS_TYPE_HEADER_EXACT = frozenset(
    {
        "terminart",
        "typ procesu",
        "process_type",
        "process type",
        "typ uslugi",
        "typ usługi",
        "leistungstyp",
    }
)

# Closed v2 cell mapping. Additional Terminart strings need an explicit decision.
_TELEDERM_CELL_VALUES = frozenset(
    {
        "hautarzt-videosprechstunde mit professioneller bilddokumentation",
    }
)


def _format_xlsx_cell_text(value) -> str:
    """Render an XLSX cell as display text (Excel numbers → text without ``.0``)."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _normalize_imported_postal_code(value) -> str | None:
    """Postleitzahl from Doctolib XLSX — often stored as Excel number (``17498.0``)."""
    text = _format_xlsx_cell_text(value)
    if not text:
        return None
    # Legacy rows already saved as ``15537.0`` string.
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    return text


def _normalize_header_cell(cell_value: str | None) -> str:
    if cell_value is None:
        return ""
    return str(cell_value).strip().lower().replace("  ", " ")


def _find_header_indices(row: list) -> dict[str, int]:
    """Map internal key -> column index (0-based). Row is list of cell values."""
    result: dict[str, int] = {}
    used: set[int] = set()
    for idx, raw in enumerate(row):
        normalized = _normalize_header_cell(raw)
        if not normalized:
            continue
        if normalized in PROCESS_TYPE_HEADER_EXACT:
            result["process_type"] = idx
            used.add(idx)
            break
    for idx, raw in enumerate(row):
        if idx in used:
            continue
        normalized = _normalize_header_cell(raw)
        if not normalized:
            continue
        for key, aliases in HEADER_ALIASES.items():
            if key in result:
                continue
            if any(alias in normalized or normalized in alias for alias in aliases):
                result[key] = idx
                break
    return result


def map_xlsx_process_type_cell(raw: str | None) -> tuple[str, bool]:
    """Map a v2 process-type cell to ProcessType.

    Returns (process_type, used_fallback). Empty or unknown → STANDARD + fallback.
    """
    text = (raw or "").strip()
    if not text:
        return ProcessType.STANDARD, True
    key = " ".join(text.lower().split())
    if key in _TELEDERM_CELL_VALUES or key == ProcessType.TELEDERM.lower():
        return ProcessType.TELEDERM, False
    if key == ProcessType.STANDARD.lower():
        return ProcessType.STANDARD, False
    return ProcessType.STANDARD, True


GERMAN_MONTHS = {
    "januar": 1,
    "februar": 2,
    "maerz": 3,
    "märz": 3,
    "april": 4,
    "mai": 5,
    "juni": 6,
    "juli": 7,
    "august": 8,
    "september": 9,
    "oktober": 10,
    "november": 11,
    "dezember": 12,
}


def _parse_date(value: str | None, *, default_year: int | None = None) -> date | None:
    """Parse DD.MM.YYYY, YYYY-MM-DD, DD/MM/YYYY, or textual German date."""
    if not value:
        return None
    s = str(value).strip()
    if not s:
        return None
    # Doctolib exports may include age suffix: "4.07.1996 (30 Jahre)".
    s = re.sub(r"\s*\([^)]*\)\s*$", "", s).strip()
    # Keep only the likely date prefix if extra text remains.
    s = re.sub(r"^(\d{1,2}[./]\d{1,2}[./]\d{2,4}).*$", r"\1", s)
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue

    # e.g. "Dienstag, 30. Dezember" or "30. Dezember 2026"
    normalized = s.lower().replace(",", " ")
    normalized = re.sub(r"\s+", " ", normalized).strip()
    m = re.search(r"(\d{1,2})\.\s*([a-zA-ZäöüÄÖÜß]+)(?:\s+(\d{4}))?", normalized)
    if m:
        day = int(m.group(1))
        month_raw = m.group(2).lower()
        month_raw_ascii = (
            month_raw.replace("ä", "ae")
            .replace("ö", "oe")
            .replace("ü", "ue")
            .replace("ß", "ss")
        )
        month = GERMAN_MONTHS.get(month_raw) or GERMAN_MONTHS.get(month_raw_ascii)
        if month:
            year = (
                int(m.group(3)) if m.group(3) else (default_year or timezone.now().year)
            )
            try:
                return date(year, month, day)
            except ValueError:
                return None

    return None


def _parse_time(value: str | None) -> time | None:
    """Parse HH:MM or HH:MM:SS."""
    if not value:
        return None
    s = str(value).strip()
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


def _split_full_name(full: str) -> tuple[str, str]:
    """
    Split Doctolib `Patient:in` full name into (first_name, last_name).

    Typical input is `Frau/Herr LAST_NAME FIRST_NAME` and may include
    non-name symbols like '@' in the same cell.
    """
    cleaned = re.sub(r"[^0-9A-Za-zÄÖÜäöüßÀ-ÿ'\\-\\s]+", " ", full or "")
    parts = [p for p in cleaned.strip().split() if p]
    if not parts:
        return ("", "")

    titles = {"frau", "herr", "mr", "mrs", "ms", "dr", "prof"}
    while parts and parts[0].lower().strip(".") in titles:
        parts.pop(0)

    if not parts:
        return ("", "")
    if len(parts) == 1:
        return (parts[0], "")

    # Doctolib exports are usually: LAST_NAME FIRST_NAME [MIDDLE...]
    last_name = parts[0]
    first_name = " ".join(parts[1:])
    return (first_name, last_name)


def _title_case_name(value: str) -> str:
    from apps.reception.patient_identity import normalize_patient_name_for_storage

    return normalize_patient_name_for_storage(value)


def _normalize_site_name(name: str) -> str:
    value = unicodedata.normalize("NFKD", name)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.lower().replace("ß", "ss")
    value = re.sub(r"^standort\s+", "", value).strip()
    value = re.sub(r"[^a-z0-9]+", "", value)
    return value


def _cleanup_clinic_name(value: str) -> str:
    """
    Remove date fragments accidentally appended to clinic name, e.g.
    'Kreutzigerstraße Freitag, 6. März' -> 'Kreutzigerstraße'.
    """
    cleaned = re.sub(r"\s+", " ", value).strip(" ,;-")
    weekday_pattern = (
        r"(montag|dienstag|mittwoch|donnerstag|freitag|samstag|sonntag|"
        r"monday|tuesday|wednesday|thursday|friday|saturday|sunday)"
    )
    trailing_date_pattern = (
        rf"\s+{weekday_pattern},?\s+\d{{1,2}}\.\s*[a-zA-ZäöüÄÖÜß]+(?:\s+\d{{4}})?$"
    )
    cleaned = re.sub(trailing_date_pattern, "", cleaned, flags=re.IGNORECASE).strip(
        " ,;-"
    )
    return cleaned


def _extract_file_metadata(rows: list[list]) -> tuple[date, str]:
    """Extract (queue_date, clinic_name) from top rows."""
    year_hint = timezone.now().year
    queue_date: date | None = None
    clinic_name: str | None = None

    for row in rows[:8]:
        for idx, raw_cell in enumerate(row):
            if raw_cell is None:
                continue
            cell = str(raw_cell).strip()
            if not cell:
                continue

            if queue_date is None:
                parsed = _parse_date(cell, default_year=year_hint)
                if parsed:
                    queue_date = parsed

            if clinic_name is None:
                m = re.search(
                    r"(?:standort|clinic)\s*:?\s*(.+)$", cell, flags=re.IGNORECASE
                )
                if m:
                    clinic_name = _cleanup_clinic_name(m.group(1).strip())
                elif cell.lower() == "standort" and idx + 1 < len(row) and row[idx + 1]:
                    clinic_name = _cleanup_clinic_name(str(row[idx + 1]).strip())

    if queue_date is None:
        raise XlsxImportFailure(
            XlsxImportErrorCode.MISSING_IMPORT_DATE,
            "Nie znaleziono daty kolejki w nagłówku pliku.",
        )
    if not clinic_name:
        raise XlsxImportFailure(
            XlsxImportErrorCode.MISSING_CLINIC_NAME,
            "Nie znaleziono nazwy placówki (Standort) w nagłówku pliku.",
        )
    return queue_date, clinic_name


def _resolve_clinic_site(clinic_name: str) -> ClinicSite:
    target = _normalize_site_name(clinic_name)
    if not target:
        raise XlsxImportFailure(
            XlsxImportErrorCode.UNKNOWN_CLINIC,
            f"Nieprawidłowa nazwa placówki w pliku: {clinic_name!r}",
        )

    all_sites = list(ClinicSite.objects.all())
    matches = [site for site in all_sites if _normalize_site_name(site.name) == target]
    if not matches:
        matches = [
            site for site in all_sites if target in _normalize_site_name(site.name)
        ]
    if len(matches) != 1:
        raise XlsxImportFailure(
            XlsxImportErrorCode.UNKNOWN_CLINIC,
            f"Nie można jednoznacznie dopasować placówki z pliku: {clinic_name!r}",
        )
    return matches[0]


@dataclass
class NormalizedRow:
    row_number: int
    first_name: str
    last_name: str
    date_of_birth: date
    phone: str
    email: str
    appointment_time: time | None = None
    street: str | None = None
    postal_code: str | None = None
    city: str | None = None
    process_type: str = ProcessType.STANDARD
    process_type_fallback: bool = False


def _normalize_row(
    row_index: int,
    row: list,
    header_indices: dict[str, int],
) -> NormalizedRow | None:
    """Convert a data row to NormalizedRow. Returns None if row is empty."""

    def _cell(key: str) -> str:
        idx = header_indices.get(key, -1)
        if idx < 0 or idx >= len(row):
            return ""
        return _format_xlsx_cell_text(row[idx])

    first_name = _cell("first_name")
    last_name = _cell("last_name")
    if not first_name and not last_name:
        full = _cell("full_name")
        if full:
            first_name, last_name = _split_full_name(full)
    first_name = _title_case_name(first_name)
    last_name = _title_case_name(last_name)
    if not first_name and not last_name:
        return None
    if not first_name or not last_name:
        raise XlsxImportFailure(
            XlsxImportErrorCode.MISSING_REQUIRED_FIELD,
            f"Row {row_index}: missing first_name or last_name",
        )

    dob_idx = header_indices.get("date_of_birth", -1)
    raw_dob = row[dob_idx] if 0 <= dob_idx < len(row) else None
    dob: date | None
    if isinstance(raw_dob, datetime):
        dob = raw_dob.date()
    elif isinstance(raw_dob, date):
        dob = raw_dob
    else:
        dob_raw = str(raw_dob).strip() if raw_dob is not None else ""
        dob = _parse_date(dob_raw) if dob_raw else None
    if not dob:
        raise XlsxImportFailure(
            XlsxImportErrorCode.INVALID_DATE_OF_BIRTH,
            f"Row {row_index}: invalid or missing date of birth: {raw_dob!r}",
        )

    phone_raw = _cell("phone")
    phone = normalize_phone_for_patient_storage(phone_raw)
    if not phone or len(phone) < 7:
        raise XlsxImportFailure(
            XlsxImportErrorCode.INVALID_PHONE,
            f"Row {row_index}: invalid or missing phone: {phone_raw!r}",
        )

    email = normalize_email_for_storage(_cell("email"))
    if not email:
        raise XlsxImportFailure(
            XlsxImportErrorCode.MISSING_REQUIRED_FIELD,
            f"Row {row_index}: missing email",
        )

    time_raw = _cell("appointment_time")
    appointment_time = _parse_time(time_raw) if time_raw else None
    street = _cell("address") or None
    postal_code = _normalize_imported_postal_code(
        row[header_indices["postal_code"]]
        if "postal_code" in header_indices and header_indices["postal_code"] < len(row)
        else None
    )
    city = _cell("city") or None
    has_process_type_column = "process_type" in header_indices
    if has_process_type_column:
        process_type, process_type_fallback = map_xlsx_process_type_cell(
            _cell("process_type")
        )
    else:
        process_type, process_type_fallback = ProcessType.STANDARD, False

    return NormalizedRow(
        row_number=row_index,
        first_name=first_name,
        last_name=last_name,
        date_of_birth=dob,
        phone=phone,
        email=email,
        appointment_time=appointment_time,
        street=street,
        postal_code=postal_code,
        city=city,
        process_type=process_type,
        process_type_fallback=process_type_fallback,
    )


def _sync_patient_address_from_import_row(
    patient: Patient,
    norm: NormalizedRow,
) -> None:
    """Persist address fields from XLSX when present (Doctolib Anschrift / PLZ / Ort)."""
    update_fields: list[str] = []
    for field_name, raw in (
        ("street", norm.street),
        ("postal_code", norm.postal_code),
        ("city", norm.city),
    ):
        text = (raw or "").strip()
        if text:
            setattr(patient, field_name, text)
            update_fields.append(field_name)
    if update_fields:
        update_fields.append("updated_at")
        patient.save(update_fields=update_fields)


def _validate_headers(header_indices: dict[str, int]) -> None:
    """Require at least first_name or full_name, last_name or full_name, date_of_birth, phone, email."""
    required = ["date_of_birth", "phone", "email"]
    name_ok = (
        "first_name" in header_indices
        or "last_name" in header_indices
        or "full_name" in header_indices
    )
    if not name_ok:
        raise XlsxImportFailure(
            XlsxImportErrorCode.TEMPLATE_HEADER_INVALID,
            "Template must have first_name/last_name or full_name (Pacjent), date_of_birth, phone, email.",
        )
    for key in required:
        if key not in header_indices:
            raise XlsxImportFailure(
                XlsxImportErrorCode.TEMPLATE_HEADER_INVALID,
                f"Missing required column: {key}",
            )


def _audit_xlsx_import_finished(
    batch: PatientImportBatch,
    *,
    context_clinic_site_id: uuid.UUID | None,
    status: str,
    inserted_rows: int,
    matched_rows: int,
    skipped_already_present_count: int,
    error_rows: int,
    v2_process_type_fallback_count: int = 0,
    failure_reason: str | None = None,
) -> None:
    md: dict = {
        "batch_id": str(batch.id),
        "status": status,
        "inserted_rows": inserted_rows,
        "matched_rows": matched_rows,
        "skipped_already_present_count": skipped_already_present_count,
        "error_rows": error_rows,
        "v2_process_type_fallback_count": v2_process_type_fallback_count,
    }
    if failure_reason:
        md["failure_reason"] = failure_reason
    create_audit_event(
        event_type="PATIENT_XLSX_IMPORT_FINISHED",
        actor_user_id=batch.created_by_user_id,
        context_clinic_site_id=context_clinic_site_id,
        metadata=md,
    )


def _try_record_import_batch_finished(
    batch: PatientImportBatch,
    *,
    result: str,
) -> None:
    try:
        record_import_batch_finished(
            result=result,
            started_at=batch.created_at,
            finished_at=batch.finished_at,
        )
    except Exception:
        logger.exception(
            "record_import_batch_finished failed after batch %s (result=%s)",
            batch.id,
            result,
        )


def _finalize_patient_xlsx_import_batch_failed(
    batch: PatientImportBatch,
    *,
    context_clinic_site_id: uuid.UUID | None,
    error_code: str,
    error_message: str,
) -> None:
    """Mark batch failed, persist row-1 error, audit, and record import metrics once."""
    batch.status = ImportStatus.FAILED
    batch.finished_at = timezone.now()
    batch.save(update_fields=["status", "finished_at"])
    PatientImportError.objects.create(
        batch=batch,
        row_number=1,
        error_code=error_code,
        error_message=error_message,
        raw_row=None,
    )
    _audit_xlsx_import_finished(
        batch,
        context_clinic_site_id=context_clinic_site_id,
        status=ImportStatus.FAILED.value,  # type: ignore[attr-defined]
        inserted_rows=0,
        matched_rows=0,
        skipped_already_present_count=0,
        error_rows=0,
        failure_reason=error_message,
    )
    _try_record_import_batch_finished(batch, result="failed")


def process_patient_xlsx_import_batch(
    *,
    batch_id: uuid.UUID,
    stored_file_path: str,
) -> None:
    """
    Read XLSX from stored_file_path, validate template, normalize rows,
    upsert patients and create queue entries. Updates batch counts and status.
    """
    from openpyxl import load_workbook

    batch = PatientImportBatch.objects.get(id=batch_id)

    inserted = 0
    matched = 0
    skipped_already_present = 0
    v2_process_type_fallback = 0
    errors_count = 0
    seen_identity: set[tuple[str, str, str, date, str]] = set()
    header_indices: dict[str, int] = {}
    daily_queue_id: uuid.UUID | None = None
    queue_date: date | None = None
    clinic_site_id: uuid.UUID | None = None
    consulting_room_id: uuid.UUID | None = None
    shift_code: str | None = None
    created_by_user_id = batch.created_by_user_id
    header_row_no: int | None = None

    try:
        wb = load_workbook(stored_file_path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise XlsxImportFailure(
                XlsxImportErrorCode.TEMPLATE_HEADER_INVALID,
                "Workbook has no active sheet.",
            )

        materialized_rows = [
            list(r) if r else [] for r in ws.iter_rows(values_only=True)
        ]
        queue_date, clinic_name = _extract_file_metadata(materialized_rows)
        clinic_site = _resolve_clinic_site(clinic_name)
        clinic_site_id = clinic_site.id
        if clinic_site.pdf_import_default_consulting_room_id is None:
            raise XlsxImportFailure(
                XlsxImportErrorCode.UNKNOWN_CLINIC,
                "Placówka z pliku nie ma ustawionego domyślnego gabinetu dla importów.",
            )
        consulting_room_id = clinic_site.pdf_import_default_consulting_room_id
        shift_code = clinic_site.pdf_import_shift_code

        for row_no, row_list in enumerate(materialized_rows, start=1):
            if header_row_no is None:
                candidate = _find_header_indices(row_list)
                if (
                    "phone" in candidate
                    and "email" in candidate
                    and (
                        "full_name" in candidate
                        or "first_name" in candidate
                        or "last_name" in candidate
                    )
                ):
                    header_indices = candidate
                    _validate_headers(header_indices)
                    header_row_no = row_no
                    batch.total_rows = max(0, len(materialized_rows) - header_row_no)
                    batch.save(update_fields=["total_rows"])
                continue

            if row_no <= header_row_no:
                continue

            if not any(c is not None and str(c).strip() for c in row_list):
                continue

            try:
                norm = _normalize_row(row_no, row_list, header_indices)
                if norm is None:
                    continue
            except XlsxImportFailure as e:
                errors_count += 1
                PatientImportError.objects.create(
                    batch=batch,
                    row_number=row_no,
                    error_code=e.error_code,
                    error_message=str(e),
                    raw_row=dict(zip(range(len(row_list)), row_list)),
                )
                continue
            except DomainError as e:
                errors_count += 1
                PatientImportError.objects.create(
                    batch=batch,
                    row_number=row_no,
                    error_code=XlsxImportErrorCode.INVALID_ROW_FORMAT,
                    error_message=str(e),
                    raw_row=dict(zip(range(len(row_list)), row_list)),
                )
                continue

            identity_key = (
                *patient_identity_key(
                    first_name=norm.first_name,
                    last_name=norm.last_name,
                    phone=norm.phone,
                    date_of_birth=norm.date_of_birth,
                ),
                norm.process_type,
            )
            if identity_key in seen_identity:
                errors_count += 1
                PatientImportError.objects.create(
                    batch=batch,
                    row_number=norm.row_number,
                    error_code=XlsxImportErrorCode.DUPLICATE_IN_FILE,
                    error_message=domain_message(
                        "other.domain.import_duplicate_identity_in_file",
                        first_name=norm.first_name,
                        last_name=norm.last_name,
                        phone=norm.phone,
                        date_of_birth=norm.date_of_birth.isoformat(),
                    ),
                    raw_row={
                        "first_name": norm.first_name,
                        "last_name": norm.last_name,
                    },
                )
                continue
            seen_identity.add(identity_key)

            if norm.process_type_fallback:
                v2_process_type_fallback += 1

            existing_active = find_patient_for_import(
                first_name=norm.first_name,
                last_name=norm.last_name,
                phone=norm.phone,
                date_of_birth=norm.date_of_birth,
            )
            reused_existing = False
            if existing_active is not None:
                patient = existing_active
                reused_existing = True
            else:
                if stale_anonymized_patient_blocks_phone(phone=norm.phone):
                    errors_count += 1
                    PatientImportError.objects.create(
                        batch=batch,
                        row_number=norm.row_number,
                        error_code=XlsxImportErrorCode.PATIENT_ANONYMIZED_NEW_RECORD,
                        error_message=domain_message(
                            "other.domain.import_patient_anonymized_same_phone",
                        ),
                        raw_row={
                            "first_name": norm.first_name,
                            "last_name": norm.last_name,
                        },
                    )
                    continue
                try:
                    validate_patient_names_for_import(
                        first_name=norm.first_name,
                        last_name=norm.last_name,
                    )
                    patient = create_or_update_patient_manual(
                        first_name=norm.first_name,
                        last_name=norm.last_name,
                        date_of_birth=norm.date_of_birth,
                        phone=norm.phone,
                        email=norm.email,
                        created_or_updated_by_user_id=created_by_user_id,
                        doctolib_patient_id=None,
                        patient_id=None,
                    )
                except Exception as e:
                    errors_count += 1
                    PatientImportError.objects.create(
                        batch=batch,
                        row_number=norm.row_number,
                        error_code=XlsxImportErrorCode.INVALID_ROW_FORMAT,
                        error_message=str(e),
                        raw_row={
                            "first_name": norm.first_name,
                            "last_name": norm.last_name,
                        },
                    )
                    continue

            _sync_patient_address_from_import_row(patient, norm)

            if daily_queue_id is None:
                queue = DailyQueue.objects.filter(
                    queue_date=queue_date,
                    clinic_site_id=clinic_site_id,
                    consulting_room_id=consulting_room_id,
                    shift_code=shift_code,
                ).first()
                if queue is None:
                    queue = create_daily_queue(
                        queue_date=queue_date,
                        clinic_site_id=clinic_site_id,
                        consulting_room_id=consulting_room_id,
                        assigned_doctor_id=None,
                        shift_code=shift_code,
                        created_by_user_id=created_by_user_id,
                        source=QueueSource.IMPORT,
                    )
                daily_queue_id = queue.id

            appointment_dt = None
            if norm.appointment_time and daily_queue_id:
                appointment_dt = timezone.make_aware(
                    timezone.datetime.combine(queue_date, norm.appointment_time),
                    timezone.get_current_timezone(),
                )

            try:
                already_present = active_queue_entry_for_process_exists(
                    daily_queue_id=daily_queue_id,
                    patient_id=patient.id,
                    process_type=norm.process_type,
                )
                if already_present:
                    skipped_already_present += 1
                    continue

                create_queue_entry(
                    daily_queue_id=daily_queue_id,
                    patient_id=patient.id,
                    created_by_user_id=created_by_user_id,
                    appointment_time=appointment_dt,
                    visit_external_id=None,
                    notes=None,
                    process_type=norm.process_type,
                )
                if reused_existing:
                    matched += 1
                else:
                    inserted += 1
            except Exception as e:
                errors_count += 1
                PatientImportError.objects.create(
                    batch=batch,
                    row_number=norm.row_number,
                    error_code=XlsxImportErrorCode.DUPLICATE_VISIT,
                    error_message=str(e),
                    raw_row={
                        "first_name": norm.first_name,
                        "last_name": norm.last_name,
                    },
                )

        if header_row_no is None:
            raise XlsxImportFailure(
                XlsxImportErrorCode.TEMPLATE_HEADER_INVALID,
                "Nie znaleziono wiersza nagłówków z kolumnami pacjentów.",
            )

        wb.close()
    except XlsxImportFailure as e:
        _finalize_patient_xlsx_import_batch_failed(
            batch,
            context_clinic_site_id=clinic_site_id,
            error_code=e.error_code,
            error_message=str(e),
        )
        return
    except DomainError as e:
        _finalize_patient_xlsx_import_batch_failed(
            batch,
            context_clinic_site_id=clinic_site_id,
            error_code=XlsxImportErrorCode.TEMPLATE_HEADER_INVALID,
            error_message=str(e),
        )
        return
    except Exception as e:
        _finalize_patient_xlsx_import_batch_failed(
            batch,
            context_clinic_site_id=clinic_site_id,
            error_code=XlsxImportErrorCode.INVALID_ROW_FORMAT,
            error_message=str(e),
        )
        return

    batch.inserted_rows = inserted
    batch.matched_rows = matched
    batch.skipped_already_present_count = skipped_already_present
    batch.error_rows = errors_count
    batch.status = (
        ImportStatus.COMPLETED
        if errors_count == 0
        else ImportStatus.COMPLETED_WITH_ERRORS
    )
    batch.finished_at = timezone.now()
    batch.save(
        update_fields=[
            "inserted_rows",
            "matched_rows",
            "skipped_already_present_count",
            "error_rows",
            "status",
            "finished_at",
        ]
    )
    _audit_xlsx_import_finished(
        batch,
        context_clinic_site_id=clinic_site_id,
        status=batch.status,
        inserted_rows=inserted,
        matched_rows=matched,
        skipped_already_present_count=skipped_already_present,
        error_rows=errors_count,
        v2_process_type_fallback_count=v2_process_type_fallback,
    )
    _try_record_import_batch_finished(
        batch,
        result="completed" if errors_count == 0 else "completed_with_errors",
    )


def _store_uploaded_xlsx(uploaded_file) -> tuple[Path, str]:
    """Save uploaded file under MEDIA_ROOT/imports/patients_xlsx/, return (path, sha256)."""
    target_dir = Path(settings.MEDIA_ROOT) / "imports" / "patients_xlsx"
    target_dir.mkdir(parents=True, exist_ok=True)
    suffix = Path(uploaded_file.name or "import.xlsx").suffix or ".xlsx"
    if suffix.lower() != ".xlsx":
        suffix = ".xlsx"
    content = uploaded_file.read()
    sha = hashlib.sha256(content).hexdigest()
    safe_name = f"{sha[:16]}{suffix}"
    path = target_dir / safe_name
    path.write_bytes(content)
    return path, sha


def enqueue_patient_xlsx_import(
    *,
    uploaded_file,
    created_by_user,
) -> PatientImportBatch:
    """Store file, create batch, enqueue background task. Returns the created batch."""
    path, sha256_hex = _store_uploaded_xlsx(uploaded_file)
    source_name = uploaded_file.name or "import.xlsx"

    batch = PatientImportBatch.objects.create(
        source_file_name=source_name,
        source_file_sha256=sha256_hex,
        import_type=ImportType.DAILY_FILE_IMPORT,
        source_system=ImportSourceSystem.OTHER,
        status=ImportStatus.PROCESSING,
        total_rows=0,
        inserted_rows=0,
        matched_rows=0,
        error_rows=0,
        created_by_user=created_by_user,
    )

    from apps.reception.tasks import run_patient_xlsx_import

    run_patient_xlsx_import.enqueue(
        str(batch.id),
        str(path),
    )
    create_audit_event(
        event_type="PATIENT_XLSX_IMPORT_ENQUEUED",
        actor_user_id=batch.created_by_user_id,
        metadata={
            "batch_id": str(batch.id),
            "source_file_name": batch.source_file_name,
            "source_file_sha256": batch.source_file_sha256,
        },
    )
    return batch
